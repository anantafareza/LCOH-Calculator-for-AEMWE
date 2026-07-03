#!/usr/bin/env python3
"""
LCOH Calculator for AEMWE  |  Ananta Fareza
PySide6 GUI — Requirements: pip install PySide6 matplotlib openpyxl numpy
"""

import sys, csv, copy, re, os, json
import numpy as np

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout,
    QHBoxLayout, QFormLayout, QLabel, QLineEdit, QPushButton,
    QTableWidget, QTableWidgetItem, QGroupBox, QScrollArea,
    QDoubleSpinBox, QSpinBox, QComboBox, QSplitter, QFileDialog,
    QMessageBox, QSlider, QHeaderView, QSizePolicy, QFrame,
    QAbstractItemView, QCheckBox, QStatusBar, QStyledItemDelegate,
    QFileIconProvider, QTextBrowser,
)
from PySide6.QtCore import Qt, QTimer, QFileInfo
from PySide6.QtGui import QFont, QColor, QIcon, QPalette

import matplotlib
import mplcursors
matplotlib.rcParams['mathtext.default'] = 'regular'
matplotlib.rcParams['figure.facecolor'] = '#EEEEE8'
matplotlib.rcParams['axes.facecolor']   = '#FFFFFF'
matplotlib.rcParams['axes.edgecolor']   = '#B0AFA8'
matplotlib.rcParams['axes.labelcolor']  = '#1A1A18'
matplotlib.rcParams['xtick.color']      = '#444444'
matplotlib.rcParams['ytick.color']      = '#444444'
matplotlib.rcParams['text.color']       = '#1A1A18'
matplotlib.rcParams['grid.color']       = '#C8C8C0'
matplotlib.rcParams['legend.edgecolor'] = '#C0BFB8'
matplotlib.rcParams['legend.framealpha']= 0.92
matplotlib.use("QtAgg")
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

try:
    import openpyxl; HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── DARK PALETTE ─────────────────────────────────────────────────────────────

def make_palette():
    pal = QPalette()
    BG  = QColor("#EEEEE8"); BGA = QColor("#F5F5F0"); TEXT = QColor("#1A1A18")
    pal.setColor(QPalette.Window,          BG);   pal.setColor(QPalette.WindowText,      TEXT)
    pal.setColor(QPalette.Base,            QColor("#FFFFFF")); pal.setColor(QPalette.AlternateBase, BGA)
    pal.setColor(QPalette.Text,            TEXT);  pal.setColor(QPalette.Button,          QColor("#E4E3DC"))
    pal.setColor(QPalette.ButtonText,      TEXT);  pal.setColor(QPalette.ToolTipBase,     QColor("#FFFEF0"))
    pal.setColor(QPalette.ToolTipText,     TEXT);  pal.setColor(QPalette.Highlight,       QColor("#2D5A8E"))
    pal.setColor(QPalette.HighlightedText, QColor("#FFFFFF"))
    pal.setColor(QPalette.BrightText,      QColor("#B03030"))
    pal.setColor(QPalette.Link,            QColor("#2D5A8E"))
    pal.setColor(QPalette.Mid,             QColor("#C0BFB8"))
    pal.setColor(QPalette.Dark,            QColor("#A0A09A"))
    pal.setColor(QPalette.Shadow,          QColor("#808080"))
    return pal


# ── STYLE CONSTANTS ───────────────────────────────────────────────────────────

BTN_PRIMARY = (
    "QPushButton { background-color: #2D5A8E; color: white; padding: 7px 18px;"
    " font-weight: bold; border-radius: 4px; border: none; }"
    "QPushButton:hover  { background-color: #3B6BA0; }"
    "QPushButton:pressed{ background-color: #1E4070; }"
)
BTN_SMALL = (
    "QPushButton { background-color: #2D5A8E; color: white; padding: 3px 10px;"
    " font-weight: bold; border-radius: 3px; border: none; font-size: 10px; }"
    "QPushButton:hover { background-color: #3B6BA0; }"
)
GRP_STYLE = (
    "QGroupBox { font-weight: bold; font-size: 12px; margin-top: 12px;"
    " border: 1px solid #555555; border-radius: 4px; padding-top: 10px; }"
    "QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 4px;"
    " color: #555550; }"
)
BLUE_HEX = "#2D5A8E"; RED_HEX = "#B03030"

def _ann_offset(sel):
    """Smart annotation offset — flips below/left near axes edges to prevent clipping."""
    try:
        ax=sel.artist.axes; xl=ax.get_xlim(); yl=ax.get_ylim()
        xf=(sel.target[0]-xl[0])/(xl[1]-xl[0]+1e-12)
        yf=(sel.target[1]-yl[0])/(yl[1]-yl[0]+1e-12)
        return (-70 if xf>0.85 else 0, -34 if yf>0.78 else 26)
    except Exception: return (0, 26)

SYNTHESIS_ROUTES = [
    "Electrodeposition", "Coprecipitation", "Thermal reduction",
    "Sol-gel", "Hydrothermal synthesis",
]


# ── CHEMISTRY HELPERS ─────────────────────────────────────────────────────────

ATOMIC_MASSES = {
    "H": 1.008,  "He": 4.003,  "Li": 6.941,  "Be": 9.012,  "B": 10.811,
    "C": 12.011, "N": 14.007,  "O": 15.999,  "F": 18.998,  "Na": 22.990,
    "Mg": 24.305,"Al": 26.982, "Si": 28.086, "P": 30.974,  "S": 32.065,
    "Cl": 35.453,"K": 39.098,  "Ca": 40.078, "Sc": 44.956, "Ti": 47.867,
    "V": 50.942, "Cr": 51.996, "Mn": 54.938, "Fe": 55.845, "Co": 58.933,
    "Ni": 58.693,"Cu": 63.546, "Zn": 65.380, "Ga": 69.723, "As": 74.922,
    "Se": 78.960,"Mo": 95.960, "Ru": 101.07, "Pd": 106.42, "Ag": 107.87,
    "Sn": 118.71,"I": 126.90,  "W": 183.84,  "Ir": 192.22, "Pt": 195.08,
    "Au": 196.97,"Pb": 207.20,
}

def parse_formula_mw(formula):
    mw = 0.0
    for elem, cnt in re.findall(r'([A-Z][a-z]?)(\d*\.?\d*)', formula.strip()):
        if elem in ATOMIC_MASSES:
            mw += ATOMIC_MASSES[elem] * (float(cnt) if cnt else 1.0)
    return round(mw, 3)

_TO_SUB = str.maketrans("0123456789", "₀₁₂₃₄₅₆₇₈₉")

def chem_subscript(text):
    text = re.sub(r'\.(\d)', r'·\1', text)
    return re.sub(r'(?<=[A-Za-z])(\d+)', lambda m: m.group(0).translate(_TO_SUB), text)


class ChemFormulaDelegate(QStyledItemDelegate):
    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)
        option.text = chem_subscript(index.data(Qt.DisplayRole) or "")
    def setEditorData(self, editor, index):
        editor.setText(index.data(Qt.EditRole) or "")
    def setModelData(self, editor, model, index):
        model.setData(index, editor.text())


# ── NO-SCROLL SPINBOX ─────────────────────────────────────────────────────────

class NoScrollSpinBox(QDoubleSpinBox):
    def wheelEvent(self, event): event.ignore()

def dspin(val, mn=0.0, mx=1e9, dec=4, step=None):
    s = NoScrollSpinBox()
    s.setRange(mn, mx); s.setDecimals(dec); s.setValue(val)
    if step: s.setSingleStep(step)
    s.setMinimumWidth(120)
    return s


# ── DEFAULT DATA ──────────────────────────────────────────────────────────────

POL_CURVE = [
    (0.04,1.45955),(0.08,1.49226),(0.12,1.52374),(0.16,1.54954),(0.20,1.58316),
    (0.30,1.60877),(0.40,1.63438),(0.50,1.64931),(0.60,1.66466),(0.70,1.67672),
    (0.80,1.68950),(0.90,1.70237),(1.00,1.71111),(1.20,1.73130),(1.40,1.75046),
    (1.60,1.76866),(1.80,1.78599),(2.00,1.80250),(2.20,1.81828),(2.40,1.83338),
    (2.60,1.84790),(2.80,1.86190),(3.00,1.87545),(3.20,1.88862),(3.40,1.90149),
    (3.60,1.91414),(3.80,1.92862),(4.00,1.93523),
]

INPUTS_DEF = {
    "h2_target_kg_day":1000.0, "electrode_area_cm2":700.0, "operating_days_year":350.0,
    "faradaic_efficiency_pct":100.0, "plant_lifetime_years":30.0, "discount_rate_pct":8.0,
    "operating_temp_c":60.0,
    "hhv_kwh_per_kg":39.4, "molar_mass_h2":2.016, "electrons_per_h2":2.0,
    "faraday_constant":96485.0, "stack_non_mea_per_kw":80.0, "bop_per_kw":300.0,
    "stack_replacement_pct":40.0, "j_ref":1.0, "stack_lifetime_jref_h":20000.0,
    "degradation_exp":1.5, "elec_price_kwh":0.034, "water_price_l":0.001,
    "water_l_per_kg_h2":10.0, "maintenance_pct_capex":2.0, "labour_yr":50000.0,
}

# Human-readable labels for every INPUTS_DEF key — used in Excel export and Guide
PARAM_LABELS = {
    "h2_target_kg_day":        ("H2 production target",          "kg/day"),
    "electrode_area_cm2":      ("Electrode area per cell",        "cm2"),
    "operating_days_year":     ("Operating days per year",        "days/yr"),
    "faradaic_efficiency_pct": ("Faradaic efficiency",            "%"),
    "plant_lifetime_years":    ("Plant lifetime",                 "years"),
    "discount_rate_pct":       ("Discount rate / WACC",          "%"),
    "operating_temp_c":        ("Operating temperature",          "°C"),
    "hhv_kwh_per_kg":          ("HHV of H2",                     "kWh/kg"),
    "molar_mass_h2":           ("Molar mass H2",                  "g/mol"),
    "electrons_per_h2":        ("Electrons per H2 molecule",      "—"),
    "faraday_constant":        ("Faraday constant",               "C/mol"),
    "stack_non_mea_per_kw":    ("Stack non-MEA cost",             "$/kW_stack"),
    "bop_per_kw":              ("Balance of plant cost",          "$/kW_el"),
    "stack_replacement_pct":   ("Stack replacement cost",         "% initial stack CAPEX"),
    "j_ref":                   ("Reference current density j_ref","A/cm2"),
    "stack_lifetime_jref_h":   ("Stack lifetime at j_ref",        "hours"),
    "degradation_exp":         ("Degradation exponent",           "—"),
    "elec_price_kwh":          ("Electricity price",              "$/kWh"),
    "water_price_l":           ("Water price",                    "$/L"),
    "water_l_per_kg_h2":       ("Specific water consumption",     "L/kg H2"),
    "maintenance_pct_capex":   ("Maintenance",                    "% CAPEX/yr"),
    "labour_yr":               ("Fixed labour cost",              "$/year"),
}

BOM_OPTS = {
    "Membrane": {
        "options": {
            "PiperION [Versogen]":0.40, "Sustainion X37-50 RT [Dioxide Materials]":0.76,
            "FAA-3-50 [Fumatech]":0.26, "Custom":0.00,
        },
        "default":"PiperION [Versogen]", "quantity":1.2, "unit":"cm2/cell",
    },
    "Anode Ionomer": {
        "options": {
            "PiperION dispersion [Versogen]":0.16, "Fumion FAA-3-SOLUT-10 [Fumatech]":0.03,
            "Custom":0.00,
        },
        "default":"PiperION dispersion [Versogen]", "quantity":1.25, "unit":"mg/cm2",
    },
    "Anode Catalyst": {
        "options": {
            "Commercial IrOx [Fuel Cells ETC]":0.42,   # (#1) default changed
            "Commercial NiFeOx [MSE Supplies]":0.95,
            "Commercial NiFe-LDH [MSE Supplies]":0.77,
            "Custom":0.00,              # (#2) renamed from "In-house other"
        },
        "default":"Commercial IrOx [Fuel Cells ETC]",  # (#1) removed NiFeSB
        "quantity":2.0, "unit":"mg/cm2",
    },
    "Anode PTL": {
        "options": {
            "Ni felt [TopTiTech]":0.13, "Carbon paper H23C3 [Freudenberg]":0.19,
            "Carbon paper TGP-H-060 [Toray]":0.56, "Custom":0.00,
        },
        "default":"Ni felt [TopTiTech]", "quantity":3.4, "unit":"cm2/cell",
    },
    "Cathode Ionomer": {
        "options": {
            "PiperION dispersion [Versogen]":0.16, "Fumion FAA-3-SOLUT-10 [Fumatech]":0.03,
            "Custom":0.00,
        },
        "default":"PiperION dispersion [Versogen]", "quantity":1.1, "unit":"mg/cm2",
    },
    "Cathode Catalyst": {
        "options": {
            "Commercial Pt/C [Fuel Cells ETC]":0.145, "Commercial NiMo [Sigma-Aldrich]":0.020,
            "Commercial MoS2 [Sigma-Aldrich]":0.001,  "Custom":0.00,
        },
        "default":"Commercial Pt/C [Fuel Cells ETC]", "quantity":0.3, "unit":"mg/cm2",
    },
    "Cathode GDL": {
        "options": {
            "Carbon paper H23C3 [Freudenberg]":0.19, "Carbon paper TGP-H-060 [Toray]":0.56,
            "Ni felt [TopTiTech]":0.13, "Custom":0.00,
        },
        "default":"Carbon paper H23C3 [Freudenberg]", "quantity":2.2, "unit":"cm2/cell",
    },
    "Assembly": {
        "options":{"In-house assembly":0.10},
        "default":"In-house assembly", "quantity":1.0, "unit":"per cell",
    },
}

# (#2) Default catalyst params — empty composition/table
CAT_PARAMS_DEF = {
    "name":"", "route":"Coprecipitation", "yield_pct":70.0, "overhead_pct":50.0, "formula_mw":0.0,
}


# ── APP STATE ─────────────────────────────────────────────────────────────────

class AppState:
    def __init__(self):
        self.inputs = copy.deepcopy(INPUTS_DEF)
        self.pol = copy.deepcopy(POL_CURVE)
        # Named polarisation curves — list of {"name": str, "data": [(j,V),...]}
        # The active curve (self.pol) always mirrors curves[active_curve_idx]["data"]
        self.pol_curves = [{"name": "Default (built-in)", "data": copy.deepcopy(POL_CURVE)}]
        self.active_curve_idx = 0
        # (#2 #3) Separate anode/cathode catalyst state
        self.anode_cat_params    = copy.deepcopy(CAT_PARAMS_DEF)
        self.anode_cat_reagents  = []
        self.cathode_cat_params  = copy.deepcopy(CAT_PARAMS_DEF)
        self.cathode_cat_reagents = []
        self.bom = {}
        for comp, d in BOM_OPTS.items():
            mat = d["default"]
            self.bom[comp] = {"material":mat, "quantity":d["quantity"], "unit":d["unit"],
                              "ref_cost":d["options"][mat], "override":None}
        # Saved scenarios for sensitivity comparison
        self.scenarios = []  # list of {"name": str, "results": [...calc_all output...]}

    def eff_cost(self, comp):
        b = self.bom[comp]
        c = b["override"] if b["override"] is not None else b["ref_cost"]
        return b["quantity"] * c

    def mea_per_cm2(self):
        return sum(self.eff_cost(c) for c in self.bom)

    def calc_cat_cost(self, side="anode"):
        p    = self.anode_cat_params   if side == "anode" else self.cathode_cat_params
        regs = self.anode_cat_reagents if side == "anode" else self.cathode_cat_reagents
        fw = p["formula_mw"]; y = p["yield_pct"] / 100.0; oh = p["overhead_pct"] / 100.0
        if fw == 0 or y == 0 or not regs: return 0.0
        raw = 0.0
        for r in regs:
            if r["solvent"]: raw += r["vol"] * r["price"]
            elif fw > 0 and y > 0: raw += (r["stoich"] * r["mw"] / (fw * y)) * r["price"]
        return raw * (1 + oh) / 1000.0

    def to_dict(self):
        """Serialise full app state to a JSON-safe dict."""
        return {
            "version": 2,
            "inputs": copy.deepcopy(self.inputs),
            "pol_curves": copy.deepcopy(self.pol_curves),
            "active_curve_idx": self.active_curve_idx,
            "anode_cat_params": copy.deepcopy(self.anode_cat_params),
            "anode_cat_reagents": copy.deepcopy(self.anode_cat_reagents),
            "cathode_cat_params": copy.deepcopy(self.cathode_cat_params),
            "cathode_cat_reagents": copy.deepcopy(self.cathode_cat_reagents),
            "bom": copy.deepcopy(self.bom),
        }

    def from_dict(self, d):
        """Restore state from a dict (loaded from JSON). Returns list of warnings."""
        warnings = []
        self.inputs = {**copy.deepcopy(INPUTS_DEF), **d.get("inputs", {})}
        curves = d.get("pol_curves")
        if curves:
            self.pol_curves = curves
        else:
            # legacy v1 file — only had a single pol list
            pol = d.get("pol")
            if pol:
                self.pol_curves = [{"name": "Loaded curve", "data": pol}]
            warnings.append("Polarisation curve data migrated from legacy format.")
        self.active_curve_idx = min(d.get("active_curve_idx", 0), len(self.pol_curves) - 1)
        self.pol = [tuple(p) for p in self.pol_curves[self.active_curve_idx]["data"]]
        self.anode_cat_params    = d.get("anode_cat_params",    copy.deepcopy(CAT_PARAMS_DEF))
        self.anode_cat_reagents  = d.get("anode_cat_reagents",  [])
        self.cathode_cat_params  = d.get("cathode_cat_params",  copy.deepcopy(CAT_PARAMS_DEF))
        self.cathode_cat_reagents= d.get("cathode_cat_reagents",[])
        saved_bom = d.get("bom", {})
        for comp, d2 in BOM_OPTS.items():
            if comp in saved_bom:
                self.bom[comp] = saved_bom[comp]
            else:
                mat = d2["default"]
                self.bom[comp] = {"material":mat,"quantity":d2["quantity"],"unit":d2["unit"],
                                  "ref_cost":d2["options"][mat],"override":None}
                warnings.append(f"BOM component '{comp}' not found in file — reset to default.")
        self.scenarios = []
        return warnings


# ── CALCULATOR ────────────────────────────────────────────────────────────────

class Calculator:
    def __init__(self, inputs, mea_per_cm2, pol):
        self.i = inputs; self.mea = mea_per_cm2
        self.pol_j = np.array([p[0] for p in pol])
        self.pol_v = np.array([p[1] for p in pol])

    def crf(self):
        r = self.i["discount_rate_pct"] / 100.0; n = self.i["plant_lifetime_years"]
        return (r * (1+r)**n / ((1+r)**n - 1)) if r > 0 else 1.0/n

    @staticmethod
    def thermoneutral_voltage(T_c):
        """
        Temperature-corrected thermoneutral voltage [V].
        Linear approximation valid 20-90 °C:
          V_tn(T) = (285830 - 48.5*(T-25)) / (2*96485)
        At 25 °C → 1.4813 V (consistent with HHV = 285.83 kJ/mol).
        """
        dH = 285830.0 - 48.5 * (T_c - 25.0)   # J/mol, linear fit to NIST data
        return dH / (2.0 * 96485.0)

    def calc(self, j):
        i = self.i
        A=i["electrode_area_cm2"]; fe=i["faradaic_efficiency_pct"]/100.0
        ne=i["electrons_per_h2"]; F=i["faraday_constant"]; M=i["molar_mass_h2"]
        V=float(np.interp(j,self.pol_j,self.pol_v)); I=j*A
        N = (i["h2_target_kg_day"]/86400.0) / ((I*fe)/(ne*F)*(M/1000.0))
        P_kw = N*V*I/1000.0
        mea_c=self.mea*A*N; non_mea=i["stack_non_mea_per_kw"]*P_kw; bop=i["bop_per_kw"]*P_kw
        capex=mea_c+non_mea+bop
        t_life=i["stack_lifetime_jref_h"]*(i["j_ref"]/j)**i["degradation_exp"]
        op_h=i["operating_days_year"]*24.0
        stack_rep=(op_h/t_life)*(i["stack_replacement_pct"]/100.0)*(mea_c+non_mea)
        ann_capex=capex*self.crf()
        h2_yr=i["h2_target_kg_day"]*i["operating_days_year"]
        elec=P_kw*i["elec_price_kwh"]*op_h; water=h2_yr*i["water_l_per_kg_h2"]*i["water_price_l"]
        maint=(i["maintenance_pct_capex"]/100.0)*capex; labour=i["labour_yr"]
        opex=elec+water+maint+labour
        lcoh=(ann_capex+stack_rep+opex)/h2_yr
        # Temperature-corrected efficiency: η = V_tn(T) / V_cell
        T_c = i.get("operating_temp_c", 25.0)
        V_tn = Calculator.thermoneutral_voltage(T_c)
        eff = (V_tn / V) * 100.0
        return dict(j=j,V=V,eff=eff,N=N,P_kw=P_kw,mea_cost=mea_c,non_mea=non_mea,bop=bop,
                    capex=capex,ann_capex=ann_capex,stack_rep=stack_rep,elec=elec,water=water,
                    maint=maint,labour=labour,opex=opex,h2_yr=h2_yr,lcoh=lcoh,
                    capex_contrib=(ann_capex+stack_rep)/h2_yr, opex_contrib=opex/h2_yr,
                    V_tn=V_tn)

    def calc_all(self): return [self.calc(j) for j in self.pol_j]


# ── HELPERS ───────────────────────────────────────────────────────────────────

def make_group(title):
    g = QGroupBox(title); g.setStyleSheet(GRP_STYLE); return g

def bold_label(text):
    lbl = QLabel(text); f = QFont(); f.setBold(True); lbl.setFont(f); return lbl

def hline():
    f = QFrame(); f.setFrameShape(QFrame.HLine); f.setFrameShadow(QFrame.Sunken); return f

def inst_label(text):
    lbl = QLabel(text); lbl.setWordWrap(True)
    lbl.setStyleSheet("color: #B03030; font-size: 11px; padding: 2px 0;"); return lbl

def save_fig_dialog(parent, fig, default_name="figure"):
    path, _ = QFileDialog.getSaveFileName(parent, "Save Figure", default_name,
        "PNG Files (*.png);;EPS Files (*.eps);;SVG Files (*.svg)")
    if not path: return
    try:
        fig.savefig(path, bbox_inches="tight", dpi=300)
        QMessageBox.information(parent, "Saved", f"Figure saved:\n{path}")
    except Exception as ex:
        QMessageBox.critical(parent, "Error saving figure", str(ex))

def chart_header(parent, title, fig, fname):
    row = QHBoxLayout(); row.addWidget(bold_label(title)); row.addStretch()
    sb = QPushButton("Save Figure"); sb.setStyleSheet(BTN_SMALL)
    sb.clicked.connect(lambda: save_fig_dialog(parent, fig, fname))
    row.addWidget(sb); return row


# ── TAB 1: INPUTS ─────────────────────────────────────────────────────────────

class InputsTab(QWidget):
    def __init__(self, state):
        super().__init__(); self.state=state; self.widgets={}
        self._dirty_cb = None
        self._build()

    def set_dirty_callback(self, cb): self._dirty_cb = cb

    def _mark_dirty(self):
        if self._dirty_cb: self._dirty_cb()

    def _build(self):
        outer=QVBoxLayout(self); outer.setContentsMargins(0,0,0,0)
        scroll=QScrollArea(); scroll.setWidgetResizable(True); container=QWidget()
        lay=QVBoxLayout(container); lay.setSpacing(12); lay.setContentsMargins(16,16,16,16)
        lay.addWidget(inst_label("Adjust parameters below and click 'Apply Changes'."))

        def grp(title): g=make_group(title); f=QFormLayout(g); f.setSpacing(8); return g,f
        ga,fa=grp("A.  Plant Design Target")
        self._sp(fa,"H2 production target [kg/day]","h2_target_kg_day",1,1e6,0)
        self._sp(fa,"Electrode area per cell [cm2]","electrode_area_cm2",1,10000,1)
        self._sp(fa,"Operating days per year [days/yr]","operating_days_year",1,365,0)
        self._sp(fa,"Faradaic efficiency [%]","faradaic_efficiency_pct",50,100,1)
        lay.addWidget(ga)

        gb,fb=grp("B.  Financial Parameters")
        self._sp(fb,"Plant lifetime [years]","plant_lifetime_years",1,60,0)
        self._sp(fb,"Discount rate / WACC [%]","discount_rate_pct",0.1,30,2)
        self.crf_lbl=QLabel(); fb.addRow("Capital recovery factor [CRF]:",self.crf_lbl)
        lay.addWidget(gb)

        gc,fc=grp("C.  Electrochemical Constants")
        self._sp(fc,"Operating temperature [°C]","operating_temp_c",20,90,1)
        self._sp(fc,"HHV of H2 [kWh/kg]","hhv_kwh_per_kg",30,50,2)
        self._sp(fc,"Molar mass H2 [g/mol]","molar_mass_h2",1,10,4)
        self._sp(fc,"Electrons per H2 molecule [n]","electrons_per_h2",1,4,0)
        self._sp(fc,"Faraday constant [C/mol]","faraday_constant",90000,100000,0)
        self.vtn_lbl=QLabel(); fc.addRow("Thermoneutral voltage V_tn [V]:",self.vtn_lbl)
        self.widgets["operating_temp_c"].valueChanged.connect(self._refresh_vtn)
        self._refresh_vtn()
        lay.addWidget(gc)

        gd,fd=grp("D.  Capital Expenditure")
        self._sp(fd,"Stack non-MEA cost [$/kW_stack]","stack_non_mea_per_kw",0,2000,2)
        self._sp(fd,"Balance of plant cost [$/kW_el]","bop_per_kw",0,3000,2)
        self._sp(fd,"Stack replacement cost [% initial stack CAPEX]","stack_replacement_pct",0,100,1)
        self._sp(fd,"Reference current density j_ref [A/cm2]","j_ref",0.1,10,2)
        self._sp(fd,"Stack lifetime at j_ref [hours]","stack_lifetime_jref_h",1000,5e5,0)
        self._sp(fd,"Degradation exponent [n]","degradation_exp",0.5,5,2)
        lay.addWidget(gd)

        ge,fe=grp("E.  Operating Expenditure")
        self._sp(fe,"Electricity price [$/kWh]","elec_price_kwh",0.001,2.0,4)
        self._sp(fe,"Water price [$/L]","water_price_l",0,1.0,4)
        self._sp(fe,"Specific water consumption [L/kg H2]","water_l_per_kg_h2",1,100,1)
        self._sp(fe,"Maintenance [% of CAPEX/yr]","maintenance_pct_capex",0,20,2)
        self._sp(fe,"Fixed labour cost [$/year]","labour_yr",0,5e6,0)
        lay.addWidget(ge)

        # ── Polarisation curve ─────────────────────────────────────────────
        gf=make_group("F.  AEMWE System Polarisation Curve")
        lf=QVBoxLayout(gf); lf.setSpacing(6)
        lf.addWidget(QLabel("Pairs must be sorted by current density. Voltage is linearly interpolated."))
        self.pol_tbl=QTableWidget(); self.pol_tbl.setColumnCount(2)
        self.pol_tbl.setHorizontalHeaderLabels(["Current Density [A/cm2]","Cell Voltage [V]"])
        self.pol_tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.pol_tbl.setFixedHeight(240); self._load_pol(); lf.addWidget(self.pol_tbl)
        pb=QHBoxLayout()
        for txt,fn in [("+ Add Row",self._add_pol_row),("- Remove Selected",self._del_pol_row)]:
            b=QPushButton(txt); b.setStyleSheet(BTN_PRIMARY); b.clicked.connect(fn); pb.addWidget(b)
        pb.addStretch(); lf.addLayout(pb)
        lay.addWidget(gf)

        lay.addStretch()
        br=QHBoxLayout(); ab=QPushButton("Apply Changes"); ab.setStyleSheet(BTN_PRIMARY)
        ab.clicked.connect(self.apply); br.addStretch(); br.addWidget(ab); lay.addLayout(br)
        scroll.setWidget(container); outer.addWidget(scroll); self._refresh_crf()

    def _sp(self,form,label,key,mn,mx,dec):
        s=dspin(self.state.inputs[key],mn,mx,dec)
        s.valueChanged.connect(self._refresh_crf)
        s.valueChanged.connect(self._mark_dirty)
        self.widgets[key]=s; form.addRow(f"{label}:",s)

    def _refresh_crf(self):
        rw=self.widgets.get("discount_rate_pct"); nw=self.widgets.get("plant_lifetime_years")
        if rw and nw:
            r=rw.value()/100.0; n=nw.value()
            crf=(r*(1+r)**n/((1+r)**n-1)) if r>0 else 1/n
            self.crf_lbl.setText(f"{crf:.8f}")

    def _refresh_vtn(self):
        T=self.widgets["operating_temp_c"].value()
        vtn=Calculator.thermoneutral_voltage(T)
        self.vtn_lbl.setText(f"{vtn:.5f} V")

    def _load_pol(self):
        self.pol_tbl.setRowCount(len(self.state.pol))
        for i,(j,v) in enumerate(self.state.pol):
            self.pol_tbl.setItem(i,0,QTableWidgetItem(str(j)))
            self.pol_tbl.setItem(i,1,QTableWidgetItem(str(v)))

    def _add_pol_row(self):
        r=self.pol_tbl.rowCount(); self.pol_tbl.insertRow(r)
        for c in range(2): self.pol_tbl.setItem(r,c,QTableWidgetItem("0.0"))

    def _del_pol_row(self):
        row=self.pol_tbl.currentRow()
        if row>=0: self.pol_tbl.removeRow(row)

    def apply(self):
        for key,w in self.widgets.items(): self.state.inputs[key]=w.value()
        pol=[]
        for r in range(self.pol_tbl.rowCount()):
            try: pol.append((float(self.pol_tbl.item(r,0).text()),float(self.pol_tbl.item(r,1).text())))
            except Exception: pass
        pol.sort(key=lambda x:x[0])
        if len(pol)<2: QMessageBox.warning(self,"Too few points","Need at least 2 points."); return
        vs=[p[1] for p in pol]
        if any(vs[i]>=vs[i+1] for i in range(len(vs)-1)):
            QMessageBox.warning(self,"Non-monotonic voltage",
                "Cell voltage must increase with current density. Check your j-V data.")
            return
        self.state.pol=pol
        # Keep state.pol_curves[0] in sync (single curve)
        self.state.pol_curves[0]["data"]=pol
        if self._dirty_cb: self._dirty_cb(clean=True)
        QMessageBox.information(self,"Applied","Inputs applied. Switch to LCOH Results to recalculate.")

    def load_state(self):
        """Repopulate all widgets from state after project load."""
        for key,w in self.widgets.items():
            w.blockSignals(True); w.setValue(self.state.inputs.get(key, INPUTS_DEF.get(key,0))); w.blockSignals(False)
        self._refresh_crf(); self._refresh_vtn(); self._load_pol()


# ── TAB 2: BOM / MEA MATERIALS ────────────────────────────────────────────────

class BOMTab(QWidget):
    def __init__(self, state):
        super().__init__(); self.state=state
        self._combos={}; self._qty={}; self._ov_chk={}; self._ov_spin={}
        self._ref_lbl={}; self._eff_lbl={}
        self._dirty_cb=None
        self._build()

    def set_dirty_callback(self, cb): self._dirty_cb=cb
    def _mark_dirty(self):
        if self._dirty_cb: self._dirty_cb()

    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(16,16,16,16); lay.setSpacing(10)
        lay.addWidget(inst_label("Select materials for each MEA component. Tick 'Override' to enter a custom cost."))
        comps=list(BOM_OPTS.keys())
        self.tbl=QTableWidget(len(comps),7)
        self.tbl.setHorizontalHeaderLabels(["Component","Material","Qty","Unit","Ref Cost [$/unit]","Override Cost","Effective [$/cm2]"])
        hh=self.tbl.horizontalHeader()
        hh.setSectionResizeMode(0,QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(1,QHeaderView.Stretch)
        for c in [2,3,4,5]: hh.setSectionResizeMode(c,QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(6,QHeaderView.Stretch)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)

        for row,comp in enumerate(comps):
            d=BOM_OPTS[comp]; b=self.state.bom[comp]
            self.tbl.setItem(row,0,QTableWidgetItem(comp))
            cb=QComboBox()
            for mat in d["options"]: cb.addItem(mat)
            cb.setCurrentText(b["material"]); self._combos[comp]=cb; self.tbl.setCellWidget(row,1,cb)
            qs=NoScrollSpinBox(); qs.setRange(0.0001,10000); qs.setDecimals(4); qs.setValue(b["quantity"])
            self._qty[comp]=qs; self.tbl.setCellWidget(row,2,qs)
            self.tbl.setItem(row,3,QTableWidgetItem(b["unit"]))
            rl=QLabel(f"  {b['ref_cost']:.5f}"); self._ref_lbl[comp]=rl; self.tbl.setCellWidget(row,4,rl)
            ov_w=QWidget(); ov_l=QHBoxLayout(ov_w); ov_l.setContentsMargins(4,2,4,2)
            ov_chk=QCheckBox(); ov_chk.setChecked(b["override"] is not None)
            ov_spin=NoScrollSpinBox(); ov_spin.setRange(0,10000); ov_spin.setDecimals(6)
            ov_spin.setValue(b["override"] if b["override"] is not None else b["ref_cost"])
            ov_spin.setEnabled(b["override"] is not None)
            ov_l.addWidget(ov_chk); ov_l.addWidget(ov_spin)
            self._ov_chk[comp]=ov_chk; self._ov_spin[comp]=ov_spin; self.tbl.setCellWidget(row,5,ov_w)
            el=QLabel(); self._eff_lbl[comp]=el; self.tbl.setCellWidget(row,6,el)
            cb.currentTextChanged.connect(lambda _,c=comp: self._mat_changed(c))
            cb.currentTextChanged.connect(lambda _: self._mark_dirty())
            qs.valueChanged.connect(lambda _,c=comp: self._recompute(c))
            qs.valueChanged.connect(lambda _: self._mark_dirty())
            ov_chk.stateChanged.connect(lambda _,c=comp: self._ov_toggled(c))
            ov_chk.stateChanged.connect(lambda _: self._mark_dirty())
            ov_spin.valueChanged.connect(lambda _,c=comp: self._recompute(c))
            ov_spin.valueChanged.connect(lambda _: self._mark_dirty())
            self._recompute(comp)

        lay.addWidget(self.tbl)
        tot_row=QHBoxLayout(); tot_row.addStretch(); tot_row.addWidget(QLabel("Total MEA Cost:"))
        self.total_lbl=QLabel("—"); f=QFont(); f.setBold(True); f.setPointSize(12)
        self.total_lbl.setFont(f); self.total_lbl.setStyleSheet("color: #2D5A8E;")
        tot_row.addWidget(self.total_lbl); lay.addLayout(tot_row)
        br=QHBoxLayout(); br.addStretch()
        ab=QPushButton("Apply Changes"); ab.setStyleSheet(BTN_PRIMARY); ab.clicked.connect(self.apply)
        br.addWidget(ab); lay.addLayout(br); self._refresh_total()

    def _mat_changed(self,comp):
        mat=self._combos[comp].currentText()
        ref=BOM_OPTS[comp]["options"].get(mat,0.0)
        self._ref_lbl[comp].setText(f"  {ref:.5f}")
        if not self._ov_chk[comp].isChecked(): self._ov_spin[comp].setValue(ref)
        self._recompute(comp)

    def _ov_toggled(self,comp):
        en=self._ov_chk[comp].isChecked(); self._ov_spin[comp].setEnabled(en)
        if not en: self._ov_spin[comp].setValue(BOM_OPTS[comp]["options"].get(self._combos[comp].currentText(),0.0))
        self._recompute(comp)

    def _eff_cost(self,comp):
        qty=self._qty[comp].value()
        c=(self._ov_spin[comp].value() if self._ov_chk[comp].isChecked()
           else BOM_OPTS[comp]["options"].get(self._combos[comp].currentText(),0.0))
        return qty*c

    def _recompute(self,comp):
        self._eff_lbl[comp].setText(f"  {self._eff_cost(comp):.5f}")
        self._refresh_total()

    def _refresh_total(self):
        total=sum(self._eff_cost(c) for c in self._qty)
        if hasattr(self,"total_lbl"): self.total_lbl.setText(f"  {total:.6f}  $/cm2")

    def apply(self):
        for comp in BOM_OPTS:
            b=self.state.bom[comp]; b["material"]=self._combos[comp].currentText()
            b["quantity"]=self._qty[comp].value()
            b["ref_cost"]=BOM_OPTS[comp]["options"].get(b["material"],0.0)
            b["override"]=(self._ov_spin[comp].value() if self._ov_chk[comp].isChecked() else None)
        if self._dirty_cb: self._dirty_cb(clean=True)
        QMessageBox.information(self,"Applied","BOM applied. Switch to LCOH Results to recalculate.")

    def load_state(self):
        """Repopulate BOM widgets from state after project load."""
        for comp in BOM_OPTS:
            b=self.state.bom[comp]
            self._combos[comp].blockSignals(True)
            self._combos[comp].setCurrentText(b["material"])
            self._combos[comp].blockSignals(False)
            self._qty[comp].blockSignals(True)
            self._qty[comp].setValue(b["quantity"])
            self._qty[comp].blockSignals(False)
            ref=b["ref_cost"]; self._ref_lbl[comp].setText(f"  {ref:.5f}")
            has_ov=b["override"] is not None
            self._ov_chk[comp].blockSignals(True)
            self._ov_chk[comp].setChecked(has_ov)
            self._ov_chk[comp].blockSignals(False)
            self._ov_spin[comp].blockSignals(True)
            self._ov_spin[comp].setValue(b["override"] if has_ov else ref)
            self._ov_spin[comp].setEnabled(has_ov)
            self._ov_spin[comp].blockSignals(False)
            self._recompute(comp)
        self._refresh_total()


# ── TABS 3 & 4: CATALYST SYNTHESIS (generic — anode or cathode) ──────────────

class CatalystSynthesisTab(QWidget):
    """
    (#3) Generic catalyst synthesis tab — instantiated for both anode and cathode.
    side: "Anode" or "Cathode"
    """
    def __init__(self, state: AppState, side: str):
        super().__init__(); self.state=state; self.side=side; self._build()

    # ── state helpers ─────────────────────────────────────────────────────────
    def _p(self):  # current params dict
        return self.state.anode_cat_params if self.side=="Anode" else self.state.cathode_cat_params
    def _r(self):  # current reagents list
        return self.state.anode_cat_reagents if self.side=="Anode" else self.state.cathode_cat_reagents
    def _set_p(self,p):
        if self.side=="Anode": self.state.anode_cat_params=p
        else: self.state.cathode_cat_params=p
    def _set_r(self,r):
        if self.side=="Anode": self.state.anode_cat_reagents=r
        else: self.state.cathode_cat_reagents=r
    def _bom_comp(self): return f"{self.side} Catalyst"
    def _side_key(self): return self.side.lower()

    def _build(self):
        outer=QVBoxLayout(self); outer.setContentsMargins(0,0,0,0)
        scroll=QScrollArea(); scroll.setWidgetResizable(True); container=QWidget()
        lay=QVBoxLayout(container); lay.setSpacing(12); lay.setContentsMargins(16,16,16,16)
        lay.addWidget(inst_label(f"Enter {self.side.lower()} catalyst composition to auto-calculate the formula MW. Click 'Apply Changes' to push cost to BOM."))

        gp=make_group("A.  Catalyst Parameters"); fp=QFormLayout(gp); fp.setSpacing(8)
        self.name_edit=QLineEdit(self._p()["name"])
        self.name_edit.setPlaceholderText("e.g. Ni3Fe1")
        self.name_edit.textChanged.connect(self._composition_changed)
        fp.addRow("Catalyst composition:",self.name_edit)

        # (#5) Route dropdown instead of QLineEdit
        self.route_combo=QComboBox()
        for route in SYNTHESIS_ROUTES: self.route_combo.addItem(route)
        self.route_combo.setCurrentText(self._p().get("route","Coprecipitation"))
        fp.addRow("Synthesis route:",self.route_combo)

        self.yield_sp=dspin(self._p()["yield_pct"],1,100,1)
        self.oh_sp   =dspin(self._p()["overhead_pct"],0,500,1)
        self.fw_sp   =dspin(self._p()["formula_mw"],0,9999,3)
        fp.addRow("Synthesis yield [%]:",             self.yield_sp)
        fp.addRow("Lab overhead [% of reagent cost]:", self.oh_sp)
        fp.addRow("Product formula MW [g/mol]:",       self.fw_sp)
        lay.addWidget(gp)

        gr=make_group("B.  Reagent Solvent Table"); rl=QVBoxLayout(gr); rl.setSpacing(6)
        rl.addWidget(QLabel("Stoichiometric reagents: enter Stoich + MW, leave Vol = 0.  Solvents: set Stoich = 0 and MW = 0, enter volume in mL/g product."))
        cols=["Reagent / Chemical","Stoichiometry\n[mol/mol prod]","MW\n[g/mol]","Vol\n[mL/g prod]","Price\n[$/g or $/mL]","Source","Cost\n[$/g prod]"]
        self.rgt_tbl=QTableWidget(0,len(cols)); self.rgt_tbl.setHorizontalHeaderLabels(cols)
        hh=self.rgt_tbl.horizontalHeader(); hh.setSectionResizeMode(QHeaderView.ResizeToContents)
        hh.setStretchLastSection(False)   # no column forced-wide; each sizes to its content
        self.rgt_tbl.setItemDelegateForColumn(0,ChemFormulaDelegate(self.rgt_tbl))
        self._load_reagents(); rl.addWidget(self.rgt_tbl)
        rb=QHBoxLayout()
        for txt,fn in [("+ Row",self._add_row),("- Row",self._del_row)]:
            b=QPushButton(txt); b.setStyleSheet(BTN_PRIMARY); b.clicked.connect(fn); rb.addWidget(b)
        bc=QPushButton("Recalculate"); bc.setStyleSheet(BTN_PRIMARY); bc.clicked.connect(self._recalc)
        rb.addStretch(); rb.addWidget(bc); rl.addLayout(rb); lay.addWidget(gr)

        gs=make_group("C.  Synthesis Cost Summary"); fs=QFormLayout(gs); fs.setSpacing(8)
        self.raw_lbl=QLabel(); fs.addRow("Total raw reagent cost:",self.raw_lbl)
        self.oh_lbl =QLabel(); fs.addRow("Lab overhead:",          self.oh_lbl)
        self.tot_g_lbl=QLabel(); fs.addRow("Total synthesis cost:",self.tot_g_lbl)
        self.tot_mg_lbl=QLabel(); f=QFont(); f.setBold(True); f.setPointSize(11)
        self.tot_mg_lbl.setFont(f); self.tot_mg_lbl.setStyleSheet("color: #2D5A8E;")
        fs.addRow("Cost per mg catalyst:",self.tot_mg_lbl); lay.addWidget(gs)

        br=QHBoxLayout(); br.addStretch()
        pb=QPushButton("Apply Changes"); pb.setStyleSheet(BTN_PRIMARY); pb.clicked.connect(self.apply_and_push)
        br.addWidget(pb); lay.addLayout(br)
        scroll.setWidget(container); outer.addWidget(scroll); self._recalc()

    def _composition_changed(self,text):
        mw=parse_formula_mw(text)
        if mw>0:
            self.fw_sp.blockSignals(True); self.fw_sp.setValue(mw); self.fw_sp.blockSignals(False)

    def _load_reagents(self):
        regs=self._r(); self.rgt_tbl.setRowCount(len(regs))
        for r,reg in enumerate(regs):
            for c,v in enumerate([reg["name"],str(reg["stoich"]),str(reg["mw"]),
                                   str(reg.get("vol",0.0)),str(reg["price"]),reg["source"],""]):
                self.rgt_tbl.setItem(r,c,QTableWidgetItem(v))

    def _add_row(self):
        r=self.rgt_tbl.rowCount(); self.rgt_tbl.insertRow(r)
        for c,v in enumerate(["New reagent","1.0","100.0","0.0","1.0","Source",""]): self.rgt_tbl.setItem(r,c,QTableWidgetItem(v))

    def _del_row(self):
        row=self.rgt_tbl.currentRow()
        if row>=0: self.rgt_tbl.removeRow(row)

    def _read_rows(self):
        fw=self.fw_sp.value(); y=self.yield_sp.value()/100.0; oh=self.oh_sp.value()/100.0
        rows=[]; raw=0.0
        for r in range(self.rgt_tbl.rowCount()):
            def cell(c,rr=r): return self.rgt_tbl.item(rr,c)
            try:
                stoich=float(cell(1).text()); mw=float(cell(2).text())
                vol=float(cell(3).text()); price=float(cell(4).text())
                is_sol=(stoich==0.0 and mw==0.0)
                cost=vol*price if is_sol else ((stoich*mw/(fw*y))*price if fw>0 and y>0 else 0.0)
                raw+=cost; rows.append((cell(0).text(),stoich,mw,vol,price,cell(5).text(),cost))
            except Exception: pass
        return rows,raw,raw*oh,raw*(1+oh)

    def _recalc(self):
        rows,raw,oh,total=self._read_rows()
        for r,row in enumerate(rows):
            item=self.rgt_tbl.item(r,6)
            if item: item.setText(f"{row[6]:.5f}")
        mg=total/1000.0
        self.raw_lbl.setText(f"${raw:.5f} / g catalyst")
        self.oh_lbl.setText(f"${oh:.5f} / g  [{self.oh_sp.value():.0f}% of reagents]")
        self.tot_g_lbl.setText(f"${total:.5f} / g  |  ${total*1000:.2f} / kg")
        self.tot_mg_lbl.setText(f"${mg:.7f} / mg")
        # Live push to BOM if Custom is selected — no silent drift
        self._push_cost_to_bom_live()

    def _push_cost_to_bom_live(self):
        """Push computed catalyst cost into BOM immediately if material is Custom."""
        p=self._p(); regs=self._r()
        # Build reagent list from current table without touching state
        fw=self.fw_sp.value(); y=self.yield_sp.value()/100.0; oh=self.oh_sp.value()/100.0
        if fw==0 or y==0: return
        raw=0.0
        for r in range(self.rgt_tbl.rowCount()):
            def cell(c,rr=r): return self.rgt_tbl.item(rr,c)
            try:
                s=float(cell(1).text()); m=float(cell(2).text())
                v=float(cell(3).text()); pr=float(cell(4).text())
                is_sol=(s==0.0 and m==0.0)
                raw+= v*pr if is_sol else ((s*m/(fw*y))*pr if fw>0 and y>0 else 0.0)
            except Exception: pass
        mg_cost=raw*(1+oh)/1000.0
        comp=self._bom_comp()
        if self.state.bom[comp]["material"]=="Custom":
            self.state.bom[comp]["ref_cost"]=mg_cost

    def apply_and_push(self):
        p=self._p()
        p.update({"name":self.name_edit.text(),"route":self.route_combo.currentText(),
                  "yield_pct":self.yield_sp.value(),"overhead_pct":self.oh_sp.value(),
                  "formula_mw":self.fw_sp.value()})
        self._set_p(p)
        regs=[]
        for r in range(self.rgt_tbl.rowCount()):
            def cell(c,rr=r): return self.rgt_tbl.item(rr,c)
            try:
                s=float(cell(1).text()); m=float(cell(2).text())
                regs.append({"name":cell(0).text(),"stoich":s,"mw":m,"vol":float(cell(3).text()),
                             "price":float(cell(4).text()),"source":cell(5).text(),"solvent":(s==0 and m==0)})
            except Exception: pass
        self._set_r(regs)
        mg_cost=self.state.calc_cat_cost(self._side_key())
        comp=self._bom_comp()
        mat=self.state.bom[comp]["material"]
        if mat=="Custom":
            self.state.bom[comp]["ref_cost"]=mg_cost
            msg="Cost pushed to BOM (Custom selected)."
        else:
            msg=f"Set '{comp}' to 'Custom' in BOM / MEA Materials to link this cost."
        self._recalc()
        QMessageBox.information(self,"Applied",
            f"{self.side} catalyst synthesis cost: ${mg_cost:.7f}/mg\n\n{msg}\n\n"
            "Switch to LCOH Results to recalculate.")


# ── TAB 5: LCOH RESULTS ───────────────────────────────────────────────────────

class ResultsTab(QWidget):
    def __init__(self, state):
        super().__init__(); self.state=state; self.results=[]; self._build()

    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(12,12,12,12); lay.setSpacing(6)
        lay.addWidget(inst_label("Click 'Calculate LCOH' to run the model."))
        btn_row=QHBoxLayout()
        self.calc_btn=QPushButton("  Calculate LCOH  "); self.calc_btn.setStyleSheet(BTN_PRIMARY)
        self.exp_btn =QPushButton("Export");             self.exp_btn.setStyleSheet(BTN_PRIMARY)
        self.status  =QLabel("")
        self.calc_btn.clicked.connect(self.calculate); self.exp_btn.clicked.connect(self.export)
        btn_row.addWidget(self.calc_btn); btn_row.addWidget(self.exp_btn)
        btn_row.addSpacing(12); btn_row.addWidget(self.status); btn_row.addStretch()
        lay.addLayout(btn_row)
        self.min_lbl=QLabel(""); self.min_lbl.setStyleSheet("color:#FF1744;font-weight:bold;font-size:11px;")
        lay.addWidget(self.min_lbl)

        sp=QSplitter(Qt.Horizontal)
        lw=QWidget(); ll=QVBoxLayout(lw); ll.setContentsMargins(0,0,0,0); ll.setSpacing(4)

        COLS=[
            ("j\n[A/cm2]",".3f"),("V\n[V]",".5f"),("Efficiency\n[%HHV]",".2f"),
            ("Cells\n[#]",".0f"),("Power\n[kW]",".1f"),("CAPEX\n[$]",",.0f"),
            ("Annualised CAPEX\n[$/yr]",",.0f"),("Stack Replacement\n[$/yr]",",.0f"),
            ("OPEX\n[$/yr]",",.0f"),("H2\n[kg/yr]",",.0f"),
            ("CAPEX Contribution\n[$/kg]",".4f"),("OPEX Contribution\n[$/kg]",".4f"),
            ("LCOH\n[$/kg]",".4f"),
        ]
        self.col_fmts=[c[1] for c in COLS]
        self.tbl=QTableWidget(0,len(COLS)); self.tbl.setHorizontalHeaderLabels([c[0] for c in COLS])
        hh=self.tbl.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.ResizeToContents)
        hh.setStretchLastSection(False)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        ll.addWidget(self.tbl,stretch=1)

        # Plot settings — compact 2-column form layout so it doesn't force left panel wide (#3)
        grp_ps=make_group("Plot Settings"); ps_vlay=QVBoxLayout(grp_ps); ps_vlay.setSpacing(4)
        self.y1_min=dspin(0,0,100,1); self.y1_max=dspin(0,0,100,1)
        self.y2_min=dspin(0,0,100,1); self.y2_max=dspin(0,0,100,1)
        ps_cols=QHBoxLayout(); ps_cols.setSpacing(16)
        fl1=QFormLayout(); fl1.setSpacing(3)
        fl1.addRow("LCOH y-min [$/kg]:", self.y1_min)
        fl1.addRow("LCOH y-max [$/kg]:", self.y1_max)
        fl2=QFormLayout(); fl2.setSpacing(3)
        fl2.addRow("Breakdown y-min [$/kg]:", self.y2_min)
        fl2.addRow("Breakdown y-max [$/kg]:", self.y2_max)
        ps_cols.addLayout(fl1); ps_cols.addLayout(fl2); ps_cols.addStretch()
        ps_vlay.addLayout(ps_cols); ll.addWidget(grp_ps,stretch=0)
        for w in [self.y1_min,self.y1_max,self.y2_min,self.y2_max]: w.valueChanged.connect(self._apply_ps_live)
        sp.addWidget(lw)

        rw=QWidget(); rl=QVBoxLayout(rw); rl.setContentsMargins(0,0,0,0)
        self.fig1=Figure(); self.ax1=self.fig1.add_subplot(111)
        self.canvas1=FigureCanvas(self.fig1); self.canvas1.setSizePolicy(QSizePolicy.Expanding,QSizePolicy.Expanding)
        self.fig2=Figure(); self.ax2=self.fig2.add_subplot(111)
        self.canvas2=FigureCanvas(self.fig2); self.canvas2.setSizePolicy(QSizePolicy.Expanding,QSizePolicy.Expanding)
        for lbl_text,canvas,fig,fname in [
            ("LCOH vs. Current Density",self.canvas1,self.fig1,"LCOH_vs_j"),
            ("CAPEX / OPEX Breakdown",  self.canvas2,self.fig2,"CAPEX_OPEX_breakdown"),
        ]:
            rl.addLayout(chart_header(self,lbl_text,fig,fname)); rl.addWidget(canvas)
            if lbl_text!="CAPEX / OPEX Breakdown": rl.addWidget(hline())
        sp.addWidget(rw)
        sp.setSizes([750, 810])
        sp.setStretchFactor(0, 1)
        sp.setStretchFactor(1, 1)
        sp.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        lay.addWidget(sp,1)

    def calculate(self):
        try:
            mea=self.state.mea_per_cm2()
            self.results=Calculator(self.state.inputs,mea,self.state.pol).calc_all()
            self._fill_table(); self._plot_lcoh(); self._plot_breakdown()
            best=min(self.results,key=lambda r:r["lcoh"])
            self.min_lbl.setText(f"  Minimum LCOH:  {best['lcoh']:.4f} $/kg H2   at j = {best['j']:.2f} A/cm2     [MEA cost: {mea:.5f} $/cm2]")
            self.status.setText(f"Done — {len(self.results)} operating points calculated.")
        except Exception as ex:
            import traceback
            QMessageBox.critical(self,"Calculation Error",f"{type(ex).__name__}: {ex}\n\n{traceback.format_exc()[:800]}")

    def _fill_table(self):
        keys=["j","V","eff","N","P_kw","capex","ann_capex","stack_rep","opex","h2_yr","capex_contrib","opex_contrib","lcoh"]
        best_lcoh=min(r["lcoh"] for r in self.results)
        self.tbl.setRowCount(len(self.results))
        for row,r in enumerate(self.results):
            for col,(k,fmt) in enumerate(zip(keys,self.col_fmts)):
                item=QTableWidgetItem(f"{r[k]:{fmt}}"); item.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
                if abs(r["lcoh"]-best_lcoh)<1e-9: item.setBackground(QColor("#1565C0")); item.setForeground(QColor("#FFFFFF"))
                self.tbl.setItem(row,col,item)

    def _apply_ps_live(self):
        if not self.results: return
        for ax,canvas,sp_mn,sp_mx in [(self.ax1,self.canvas1,self.y1_min,self.y1_max),(self.ax2,self.canvas2,self.y2_min,self.y2_max)]:
            self._apply_ylim(ax,sp_mn,sp_mx); canvas.draw()

    def _apply_ylim(self,ax,sp_mn,sp_mx):
        mx=sp_mx.value()
        if mx>0: ax.set_ylim(sp_mn.value(),mx)
        else: ax.set_ylim(auto=True); ax.relim(); ax.autoscale_view(scaley=True)

    def _plot_lcoh(self):
        self.ax1.clear()
        js=[r["j"] for r in self.results]; lcohs=[r["lcoh"] for r in self.results]
        best=min(self.results,key=lambda r:r["lcoh"])
        self.ax1.plot(js,lcohs,color=BLUE_HEX,linewidth=2.0,marker="o",markersize=3.5)
        self.ax1.axvline(best["j"],color=RED_HEX,linestyle="--",linewidth=1.0,alpha=0.7)
        j_range=max(js)-min(js) if len(js)>1 else 4.0
        # (#5) position annotation above the curve at text x — never overlapping data
        ann_j = best["j"] + j_range * 0.28
        ann_j = min(ann_j, min(js) + j_range * 0.72)
        v_at_ann = float(np.interp(ann_j, js, lcohs))
        gap = max(1.2, (max(lcohs) - best["lcoh"]) * 0.22)
        xytext = (ann_j, v_at_ann + gap)
        self.ax1.annotate(f"Min: {best['lcoh']:.3f} $/kg\n@ j = {best['j']:.2f} A/cm2",
            xy=(best["j"],best["lcoh"]),xytext=xytext,fontsize=9,color=RED_HEX,fontweight="bold",ha="left",
            bbox=dict(boxstyle="round,pad=0.3",facecolor="white",alpha=0.85,edgecolor=RED_HEX,linewidth=1.0),
            arrowprops=dict(arrowstyle="->",color=RED_HEX,lw=1.0))
        self.ax1.set_xlabel("Current Density [A/cm2]",fontsize=9); self.ax1.set_ylabel("LCOH [$/kg H2]",fontsize=9)
        self.ax1.set_title("LCOH vs. Operating Current Density",fontsize=10); self.ax1.grid(True,alpha=0.25)
        self._apply_ylim(self.ax1,self.y1_min,self.y1_max)
        # (#1) remove stale cursor before creating a new one
        if hasattr(self,"_cur1") and self._cur1:
            try: self._cur1.remove()
            except Exception: pass
        self._cur1 = mplcursors.cursor(self.ax1.lines[0], hover=True)
        @self._cur1.connect("add")
        def _tt1(sel):
            sel.annotation.set_text(f"j = {sel.target[0]:.3f} A/cm2\nLCOH = {sel.target[1]:.4f} $/kg")
            sel.annotation.get_bbox_patch().set(facecolor="white", alpha=0.9, edgecolor="#1565C0", linewidth=0.8)
            sel.annotation.set_fontsize(8)
            sel.annotation.xyann = _ann_offset(sel)
            sel.annotation.set_anncoords("offset points")
        self.fig1.tight_layout(); self.canvas1.draw()

    def _plot_breakdown(self):
        self.ax2.clear(); js=[r["j"] for r in self.results]
        c1=[r["ann_capex"]/r["h2_yr"] for r in self.results]; c2=[r["stack_rep"]/r["h2_yr"] for r in self.results]
        c3=[r["elec"]/r["h2_yr"] for r in self.results];      c4=[r["water"]/r["h2_yr"] for r in self.results]
        c5=[r["maint"]/r["h2_yr"] for r in self.results];     c6=[r["labour"]/r["h2_yr"] for r in self.results]
        stacks = self.ax2.stackplot(js,c1,c2,c3,c4,c5,c6,
            labels=["Annualised CAPEX","Stack Replacement","Electricity","Water","Maintenance","Labour"],
            colors=["#1565C0","#42A5F5","#E65100","#FFA726","#2E7D32","#A5D6A7"],alpha=0.85)
        self.ax2.set_xlabel("Current Density [A/cm2]",fontsize=9); self.ax2.set_ylabel("LCOH Contribution [$/kg H2]",fontsize=9)
        self.ax2.set_title("Cost Breakdown by Operating Point",fontsize=10)
        self.ax2.legend(loc="upper right",fontsize=7,ncol=2); self.ax2.grid(True,axis="y",alpha=0.2)
        self._apply_ylim(self.ax2,self.y2_min,self.y2_max)
        # (#3) invisible marker line on the stacked total — the only reliably hoverable target
        totals=[c1[i]+c2[i]+c3[i]+c4[i]+c5[i]+c6[i] for i in range(len(js))]
        inv_line, = self.ax2.plot(js, totals, alpha=0, linewidth=0, marker="o", markersize=8)
        if hasattr(self,"_cur2") and self._cur2:
            try: self._cur2.remove()
            except Exception: pass
        self._cur2 = mplcursors.cursor(inv_line, hover=True)
        results_ref = self.results   # capture for closure
        @self._cur2.connect("add")
        def _tt2(sel):
            j_val = sel.target[0]
            r = min(results_ref, key=lambda x: abs(x["j"] - j_val))
            h = r["h2_yr"]
            ann   = r["ann_capex"] / h; sr  = r["stack_rep"] / h
            elec  = r["elec"] / h;     wat = r["water"] / h
            maint = r["maint"] / h;    lab = r["labour"] / h
            tot   = ann + sr + elec + wat + maint + lab
            sel.annotation.set_text(
                f"j = {j_val:.3f} A/cm2\n"
                f"Ann. CAPEX    = {ann:.4f} $/kg\n"
                f"Stack Rep.    = {sr:.4f} $/kg\n"
                f"Electricity   = {elec:.4f} $/kg\n"
                f"Water         = {wat:.4f} $/kg\n"
                f"Maintenance   = {maint:.4f} $/kg\n"
                f"Labour        = {lab:.4f} $/kg\n"
                f"Total         = {tot:.4f} $/kg")
            sel.annotation.get_bbox_patch().set(facecolor="white", alpha=0.9, edgecolor="#1565C0", linewidth=0.8)
            sel.annotation.set_fontsize(8)
        self.fig2.tight_layout(); self.canvas2.draw()

    def export(self):
        if not self.results: QMessageBox.warning(self,"No data","Run a calculation first."); return
        path,_=QFileDialog.getSaveFileName(self,"Export Results","LCOH_AEMWE_results",
            "CSV  (*.csv);;Tab-separated text  (*.txt);;JSON  (*.json);;Excel  (*.xlsx)")
        if not path: return
        if   path.endswith(".txt"):  self._export_txt(path)
        elif path.endswith(".json"): self._export_json(path)
        elif path.endswith(".xlsx") and HAS_OPENPYXL: self._export_xlsx(path)
        else:
            if not path.endswith(".csv"): path+=".csv"
            self._export_csv(path)

    def _hdrs(self): return ["j [A/cm2]","V [V]","Efficiency [%HHV]","Cells [#]","Power [kW]","CAPEX [$]","Annualised CAPEX [$/yr]","Stack Replacement [$/yr]","OPEX [$/yr]","H2 [kg/yr]","CAPEX Contribution [$/kg]","OPEX Contribution [$/kg]","LCOH [$/kg]"]
    def _vals(self,r): return [r["j"],r["V"],r["eff"],r["N"],r["P_kw"],r["capex"],r["ann_capex"],r["stack_rep"],r["opex"],r["h2_yr"],r["capex_contrib"],r["opex_contrib"],r["lcoh"]]
    def _export_csv(self,path):
        with open(path,"w",newline="") as f:
            w=csv.writer(f); w.writerow(self._hdrs()); [w.writerow(self._vals(r)) for r in self.results]
        QMessageBox.information(self,"Exported",f"Saved: {path}")
    def _export_txt(self,path):
        with open(path,"w") as f:
            f.write("\t".join(self._hdrs())+"\n"); [f.write("\t".join(str(v) for v in self._vals(r))+"\n") for r in self.results]
        QMessageBox.information(self,"Exported",f"Saved: {path}")
    def _export_json(self,path):
        data=[dict(zip(self._hdrs(),self._vals(r))) for r in self.results]
        with open(path,"w") as f: json.dump(data,f,indent=2)
        QMessageBox.information(self,"Exported",f"Saved: {path}")
    def _export_xlsx(self,path):
        wb=openpyxl.Workbook()

        # Sheet 1: LCOH Results
        ws=wb.active; ws.title="LCOH Results"
        ws.append(self._hdrs())
        for r in self.results: ws.append(self._vals(r))

        # Sheet 2: Inputs (human-readable labels + units)
        ws2=wb.create_sheet("Inputs")
        ws2.append(["Parameter","Value","Unit"])
        for k,v in self.state.inputs.items():
            label,unit=PARAM_LABELS.get(k,(k,""))
            ws2.append([label, v, unit])

        # Sheet 3: BOM
        ws3=wb.create_sheet("BOM / MEA Materials")
        ws3.append(["Component","Material","Quantity","Unit","Ref Cost [$/unit]","Override Cost","Effective Cost [$/cm2]"])
        for comp in BOM_OPTS:
            b=self.state.bom[comp]
            ov=b["override"] if b["override"] is not None else ""
            eff=b["quantity"]*(b["override"] if b["override"] is not None else b["ref_cost"])
            ws3.append([comp, b["material"], b["quantity"], b["unit"], b["ref_cost"], ov, eff])

        # Sheet 4: Polarisation Curve(s)
        ws4=wb.create_sheet("Polarisation Curves")
        col=1
        for curve in self.state.pol_curves:
            ws4.cell(row=1,column=col,value=curve["name"]+" — j [A/cm2]")
            ws4.cell(row=1,column=col+1,value=curve["name"]+" — V [V]")
            for row_i,(j,v) in enumerate(curve["data"],start=2):
                ws4.cell(row=row_i,column=col,value=j)
                ws4.cell(row=row_i,column=col+1,value=v)
            col+=3

        wb.save(path); QMessageBox.information(self,"Exported",f"Saved: {path}")


# ── TAB 6: SENSITIVITY ANALYSIS ───────────────────────────────────────────────

SENS_PARAMS=[
    ("elec_price_kwh","Electricity Price",0.005,0.200,0.001,4,"$/kWh"),
    ("discount_rate_pct","Discount Rate",1.0,25.0,0.5,1,"%"),
    ("stack_lifetime_jref_h","Stack Lifetime at j_ref",5000,80000,1000,0,"h"),
    ("stack_replacement_pct","Stack Replacement Cost",10.0,80.0,5.0,0,"% of stack CAPEX"),
    ("maintenance_pct_capex","Maintenance",0.5,5.0,0.5,1,"% CAPEX/yr"),
    ("labour_yr","Labour Cost",10000,500000,10000,0,"$/yr"),
]

class SensitivityTab(QWidget):
    def __init__(self, state):
        super().__init__(); self.state=state; self.base_results=[]; self._sliders={}
        self._timer=QTimer(); self._timer.setSingleShot(True); self._timer.timeout.connect(self._do_update); self._build()

    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(12,12,12,12)
        lay.addWidget(inst_label("Set base case first, then use the sliders to explore ±20% sensitivity. Use Monte Carlo Analysis below to run probabilistic uncertainty analysis."))
        MC_RANGES=[
            ("elec_price_kwh",        "Electricity [$/kWh]",    0.010, 0.080, 4),
            ("discount_rate_pct",     "Discount rate [%]",       4.0,  14.0,  1),
            ("stack_lifetime_jref_h", "Stack lifetime [h]",     8000, 40000,  0),
            ("stack_replacement_pct", "Stack replacement [%]",   20,    60,   0),
            ("maintenance_pct_capex", "Maintenance [%]",          1.0,   4.0, 1),
            ("labour_yr",             "Labour [$/yr]",          20000,100000,  0),
        ]
        sp=QSplitter(Qt.Horizontal); lw=QWidget(); ll=QVBoxLayout(lw); ll.setSpacing(8)
        base_btn=QPushButton("Set Base Case from Current Inputs"); base_btn.setStyleSheet(BTN_PRIMARY)
        base_btn.clicked.connect(self._set_base); ll.addWidget(base_btn)
        self.opt_lbl=QLabel("Base case LCOH: —")
        self.opt_lbl.setStyleSheet("font-weight:bold;color:#B03030;font-size:11px;"); ll.addWidget(self.opt_lbl)

        # ── Scenario manager ─────────────────────────────────────────────
        grp_sc=make_group("Scenarios"); gs_lay=QVBoxLayout(grp_sc); gs_lay.setSpacing(4)
        sc_btn_row=QHBoxLayout()
        save_sc_btn=QPushButton("Save current as scenario"); save_sc_btn.setStyleSheet(BTN_SMALL)
        save_sc_btn.clicked.connect(self._save_scenario)
        del_sc_btn=QPushButton("Remove selected"); del_sc_btn.setStyleSheet(BTN_SMALL)
        del_sc_btn.clicked.connect(self._del_scenario)
        sc_btn_row.addWidget(save_sc_btn); sc_btn_row.addWidget(del_sc_btn); sc_btn_row.addStretch()
        gs_lay.addLayout(sc_btn_row)
        self.sc_list=QTableWidget(0,2); self.sc_list.setHorizontalHeaderLabels(["Scenario","Min LCOH [$/kg]"])
        self.sc_list.horizontalHeader().setSectionResizeMode(0,QHeaderView.Stretch)
        self.sc_list.horizontalHeader().setSectionResizeMode(1,QHeaderView.ResizeToContents)
        self.sc_list.setFixedHeight(110); self.sc_list.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.sc_list.setSelectionBehavior(QAbstractItemView.SelectRows)
        gs_lay.addWidget(self.sc_list); ll.addWidget(grp_sc)

        # ── Sensitivity sliders ──────────────────────────────────────────
        grp_sens=make_group("Sensitivity Sliders  (±20% from base)"); gs=QVBoxLayout(grp_sens); gs.setSpacing(4)
        for i,(key,label,mn,mx,step,dec,unit) in enumerate(SENS_PARAMS):
            pw=QWidget(); pl=QVBoxLayout(pw); pl.setContentsMargins(0,2,0,2); pl.setSpacing(2)
            lrow=QHBoxLayout()
            name_lbl=QLabel(f"{label}  [{unit}]"); name_lbl.setStyleSheet("font-size:10px;font-weight:bold;")
            val_lbl=QLabel("—"); val_lbl.setMinimumWidth(80)
            base_lbl=QLabel("(base: —)"); base_lbl.setStyleSheet("color:#888880;font-size:10px;")
            lrow.addWidget(name_lbl); lrow.addStretch(); lrow.addWidget(val_lbl); lrow.addWidget(base_lbl)
            pl.addLayout(lrow)
            slider=QSlider(Qt.Horizontal); n_steps=int(round((mx-mn)/step)); slider.setRange(0,n_steps)
            cur=self.state.inputs.get(key,mn); slider.setValue(max(0,min(n_steps,int(round((cur-mn)/step)))))
            self._sliders[key]=(slider,mn,step,dec,val_lbl,base_lbl)
            val_lbl.setText(f"{mn+slider.value()*step:.{dec}f}")
            def _on(v,l=mn,s=step,d=dec,lbl=val_lbl): lbl.setText(f"{l+v*s:.{d}f}"); self._timer.start(200)
            slider.valueChanged.connect(_on); pl.addWidget(slider); gs.addWidget(pw)
            if i<len(SENS_PARAMS)-1: gs.addWidget(hline())
        ll.addWidget(grp_sens)

        # ── Monte Carlo settings ─────────────────────────────────────────
        MC_NORM_DEFS={"elec_price_kwh":(0.034,0.015),"discount_rate_pct":(8.0,2.5),
                      "stack_lifetime_jref_h":(20000,7000),"stack_replacement_pct":(40,12),
                      "maintenance_pct_capex":(2.0,0.7),"labour_yr":(50000,15000)}
        # Uniform removed — all three remaining distributions use mean/std parameterisation
        grp_mc=make_group("Monte Carlo Analysis"); gm=QFormLayout(grp_mc); gm.setSpacing(5)
        self.mc_n=QSpinBox(); self.mc_n.setRange(500,20000); self.mc_n.setSingleStep(500); self.mc_n.setValue(5000)
        gm.addRow("Samples:",self.mc_n)
        self.mc_dist=QComboBox(); self.mc_dist.addItems(["Normal","Log-normal","Gamma"])
        gm.addRow("Distribution:",self.mc_dist)
        # Column headers (update with distribution)
        hdr_w=QWidget(); hdr_l=QHBoxLayout(hdr_w); hdr_l.setContentsMargins(0,0,0,0); hdr_l.setSpacing(3)
        self.mc_col1_hdr=QLabel("Mean"); self.mc_col2_hdr=QLabel("Std dev")
        for lbl in [self.mc_col1_hdr,self.mc_col2_hdr]:
            lbl.setFixedWidth(76); lbl.setAlignment(Qt.AlignCenter)
            lbl.setStyleSheet("font-size:10px;color:#666660;font-weight:bold;")
        hdr_l.addWidget(self.mc_col1_hdr); hdr_l.addWidget(self.mc_col2_hdr); hdr_l.addStretch()
        gm.addRow("",hdr_w)
        sep_w=QWidget(); sep_l=QVBoxLayout(sep_w); sep_l.setContentsMargins(0,2,0,2); sep_l.addWidget(hline()); gm.addRow(sep_w)
        self._mc_ranges={}
        for key,label,lo_def,hi_def,dec in MC_RANGES:
            hw=QWidget(); hl=QHBoxLayout(hw); hl.setContentsMargins(0,0,0,0); hl.setSpacing(3)
            lo_sp=QDoubleSpinBox(); hi_sp=QDoubleSpinBox()
            n_lo,n_hi=MC_NORM_DEFS.get(key,(lo_def,hi_def))
            for sp2,val in [(lo_sp,n_lo),(hi_sp,n_hi)]:
                sp2.setDecimals(dec); sp2.setRange(0,1e8)
                sp2.setSingleStep(max(10**(-dec),(hi_def-lo_def)/20))
                sp2.setValue(val); sp2.setFixedWidth(76)
            hl.addWidget(lo_sp); hl.addWidget(hi_sp); hl.addStretch()
            self._mc_ranges[key]=(lo_sp,hi_sp)
            gm.addRow(label+":",hw)
        def _update_dist(dist,_n=MC_NORM_DEFS):
            # Normal, Log-normal, and Gamma all take mean/std inputs
            self.mc_col1_hdr.setText("Mean"); self.mc_col2_hdr.setText("Std dev")
            for k,(ls,hs) in self._mc_ranges.items():
                if k in _n: ls.setValue(_n[k][0]); hs.setValue(_n[k][1])
        self.mc_dist.currentTextChanged.connect(_update_dist)
        sep_w2=QWidget(); sep_l2=QVBoxLayout(sep_w2); sep_l2.setContentsMargins(0,2,0,2); sep_l2.addWidget(hline()); gm.addRow(sep_w2)
        mc_btn=QPushButton("Run Monte Carlo"); mc_btn.setStyleSheet(BTN_PRIMARY)
        mc_btn.clicked.connect(self._run_mc); gm.addRow(mc_btn)
        self.mc_status_lbl=QLabel(""); self.mc_status_lbl.setWordWrap(True)
        self.mc_status_lbl.setStyleSheet("font-weight:bold;color:#B03030;font-size:11px;")
        gm.addRow(self.mc_status_lbl)
        ll.addWidget(grp_mc); ll.addStretch()

        scroll=QScrollArea(); scroll.setWidgetResizable(True); scroll.setWidget(lw); sp.addWidget(scroll)

        # ── Right panel: LCOH vs j (top) + Histogram | Cumulative Probability (bottom) ────
        rw=QWidget(); rl=QVBoxLayout(rw); rl.setContentsMargins(0,0,0,0)
        self.fig_top=Figure(); self.ax_top=self.fig_top.add_subplot(111)
        self.canvas_top=FigureCanvas(self.fig_top)
        self.canvas_top.setSizePolicy(QSizePolicy.Expanding,QSizePolicy.Expanding)
        rl.addLayout(chart_header(self,"LCOH vs. Current Density",self.fig_top,"Sensitivity_LCOH"))
        rl.addWidget(self.canvas_top,1); rl.addWidget(hline())

        bot_sp=QSplitter(Qt.Horizontal)
        bot_sp.setHandleWidth(2)
        bot_sp.setStyleSheet("QSplitter::handle{background:#888880;}")
        hw2=QWidget(); hl2=QVBoxLayout(hw2); hl2.setContentsMargins(0,0,0,0)
        self.fig_hist=Figure(); self.ax_hist=self.fig_hist.add_subplot(111)
        self.canvas_hist=FigureCanvas(self.fig_hist)
        self.canvas_hist.setSizePolicy(QSizePolicy.Expanding,QSizePolicy.Expanding)
        hl2.addLayout(chart_header(self,"LCOH Histogram",self.fig_hist,"MC_histogram"))
        hl2.addWidget(self.canvas_hist); bot_sp.addWidget(hw2)
        cw=QWidget(); cl=QVBoxLayout(cw); cl.setContentsMargins(0,0,0,0)
        self.fig_cdf=Figure(); self.ax_cdf=self.fig_cdf.add_subplot(111)
        self.canvas_cdf=FigureCanvas(self.fig_cdf)
        self.canvas_cdf.setSizePolicy(QSizePolicy.Expanding,QSizePolicy.Expanding)
        cl.addLayout(chart_header(self,"Cumulative Probability",self.fig_cdf,"MC_CDF"))
        cl.addWidget(self.canvas_cdf); bot_sp.addWidget(cw)
        bot_sp.setSizes([500,500]); rl.addWidget(bot_sp,1)

        sp.addWidget(rw); sp.setSizes([260,810]); sp.setSizePolicy(QSizePolicy.Expanding,QSizePolicy.Expanding); lay.addWidget(sp,1)

    def _set_base(self):
        try:
            mea=self.state.mea_per_cm2(); self.base_results=Calculator(self.state.inputs,mea,self.state.pol).calc_all()
        except Exception as ex: QMessageBox.critical(self,"Error",str(ex)); return
        best=min(self.base_results,key=lambda r:r["lcoh"])
        self.opt_lbl.setText(f"Base LCOH:  {best['lcoh']:.4f} $/kg  at j = {best['j']:.2f} A/cm2")
        for key,(slider,mn,step,dec,val_lbl,base_lbl) in self._sliders.items():
            cur=self.state.inputs.get(key,mn); slider.setValue(max(0,int(round((cur-mn)/step))))
            base_lbl.setText(f"(base: {cur:.{dec}f})")
        self._do_update()

    def _save_scenario(self):
        if not self.base_results:
            QMessageBox.warning(self,"No base case","Set base case first."); return
        try:
            mea=self.state.mea_per_cm2()
            results=Calculator(self._modified_inputs(),mea,self.state.pol).calc_all()
        except Exception as ex: QMessageBox.critical(self,"Error",str(ex)); return
        name,ok=self._ask_scenario_name()
        if not ok: return
        self.state.scenarios.append({"name":name,"results":results})
        best=min(results,key=lambda r:r["lcoh"])
        row=self.sc_list.rowCount(); self.sc_list.insertRow(row)
        self.sc_list.setItem(row,0,QTableWidgetItem(name))
        self.sc_list.setItem(row,1,QTableWidgetItem(f"{best['lcoh']:.4f}"))
        self._do_update()

    def _del_scenario(self):
        row=self.sc_list.currentRow()
        if row<0: return
        self.sc_list.removeRow(row)
        self.state.scenarios.pop(row)
        self._do_update()

    def _ask_scenario_name(self, default=None):
        if default is None:
            default=f"Scenario {len(self.state.scenarios)+1}"
        result=QMessageBox(self); result.setWindowTitle("Scenario name")
        result.setText("Enter a name for this scenario:")
        le=QLineEdit(default); result.layout().addWidget(le,1,0,1,3)
        result.setStandardButtons(QMessageBox.Ok|QMessageBox.Cancel)
        if result.exec()==QMessageBox.Ok:
            return (le.text().strip() or default), True
        return default, False

    def _modified_inputs(self):
        inp=copy.deepcopy(self.state.inputs)
        for key,(slider,mn,step,dec,*_) in self._sliders.items(): inp[key]=mn+slider.value()*step
        return inp

    def _do_update(self):
        if hasattr(self,'_sens_cur') and self._sens_cur:
            try: self._sens_cur.remove()
            except Exception: pass
            self._sens_cur=None
        self.ax_top.clear()
        try: mea=self.state.mea_per_cm2(); mod_res=Calculator(self._modified_inputs(),mea,self.state.pol).calc_all()
        except Exception: self.canvas_top.draw(); return
        js=[r["j"] for r in mod_res]; data_lines=[]
        SC_COLORS=["#1B9E77","#D95F02","#7570B3","#E7298A","#66A61E","#E6AB02","#A6761D"]

        # Collect all LCOH values across every visible series to set y-axis correctly
        all_lcohs=[]

        # Saved scenarios (behind everything else)
        for i,sc in enumerate(self.state.scenarios):
            sc_js=[r["j"] for r in sc["results"]]
            sc_ls=[r["lcoh"] for r in sc["results"]]
            sl,=self.ax_top.plot(sc_js,sc_ls,
                color=SC_COLORS[i%len(SC_COLORS)],linewidth=1.6,linestyle=":",
                label=sc["name"],zorder=2)
            data_lines.append(sl); all_lcohs.extend(sc_ls)

        if self.base_results:
            bl_ls=[r["lcoh"] for r in self.base_results]
            bl,=self.ax_top.plot(js,bl_ls,color="#9E9E9E",linewidth=1.5,
                linestyle="--",alpha=0.85,label="Base case",zorder=3)
            data_lines.append(bl); all_lcohs.extend(bl_ls)

        ml_ls=[r["lcoh"] for r in mod_res]
        ml,=self.ax_top.plot(js,ml_ls,color=BLUE_HEX,linewidth=2.0,label="Modified",zorder=4)
        data_lines.append(ml); all_lcohs.extend(ml_ls)

        if mod_res:
            best=min(mod_res,key=lambda r:r["lcoh"])
            self.ax_top.axvline(best["j"],color=RED_HEX,linestyle="--",linewidth=1.0,alpha=0.7)
            self.ax_top.set_title(f"Modified  -->  Min LCOH: {best['lcoh']:.3f} $/kg  @  j = {best['j']:.2f} A/cm2",fontsize=9)

        # Y-axis: fit all series with a small margin
        if all_lcohs:
            y_lo=min(all_lcohs); y_hi=max(all_lcohs)
            margin=(y_hi-y_lo)*0.08 if y_hi>y_lo else 0.5
            self.ax_top.set_ylim(max(0,y_lo-margin), y_hi+margin)

        self.ax_top.set_xlabel("Current Density [A/cm2]",fontsize=9)
        self.ax_top.set_ylabel("LCOH [$/kg H2]",fontsize=9)
        self.ax_top.legend(fontsize=7,loc="upper right"); self.ax_top.grid(True,alpha=0.25)
        if data_lines:
            self._sens_cur=mplcursors.cursor(data_lines,hover=True)
            @self._sens_cur.connect("add")
            def _tt_top(sel):
                sel.annotation.set_text(f"j = {sel.target[0]:.3f} A/cm2\nLCOH = {sel.target[1]:.4f} $/kg")
                sel.annotation.get_bbox_patch().set(facecolor="white",alpha=0.9,edgecolor="#2D5A8E",linewidth=0.8)
                sel.annotation.set_fontsize(8); sel.annotation.xyann=_ann_offset(sel)
                sel.annotation.set_anncoords("offset points")
        self.fig_top.tight_layout(); self.canvas_top.draw()

    def _run_mc(self):
        if not self.base_results:
            QMessageBox.warning(self,"No base case","Set base case first before running Monte Carlo.")
            return
        dist=self.mc_dist.currentText()
        mea=self.state.mea_per_cm2(); n=self.mc_n.value()
        self.mc_status_lbl.setText("Running…"); QApplication.processEvents()
        # Parameter-specific minimums to keep Calculator numerically valid
        _PARAM_MIN={"stack_lifetime_jref_h":500,"elec_price_kwh":1e-4,
                    "discount_rate_pct":0.1,"stack_replacement_pct":0.1,
                    "maintenance_pct_capex":0.01,"labour_yr":1000}
        param_samples={}
        for key,(lo_sp,hi_sp) in self._mc_ranges.items():
            lo=lo_sp.value(); hi=hi_sp.value(); pmin=_PARAM_MIN.get(key,0)
            if lo<=0 or hi<=0: continue          # mean and std must both be positive
            if dist=="Gamma":
                # k = (mean/std)²,  θ = std²/mean
                # Strictly positive by definition; no clipping needed beyond pmin
                k_shape = (lo/hi)**2
                theta   = hi**2 / lo
                vals = np.random.gamma(k_shape, theta, n)
            elif dist=="Log-normal":
                # Convert (mean, std) of the resulting distribution to
                # the underlying normal parameters (μ_ln, σ_ln)
                # σ_ln = √(log(1 + (std/mean)²))
                # μ_ln = log(mean) − σ_ln²/2
                sigma_ln = np.sqrt(np.log(1.0 + (hi/lo)**2))
                mu_ln    = np.log(lo) - 0.5*sigma_ln**2
                vals = np.random.lognormal(mu_ln, sigma_ln, n)
            else:  # Normal
                vals = np.random.normal(lo, hi, n)
            vals = np.maximum(pmin, vals)
            param_samples[key] = vals
        # Run MC
        lcohs=[]
        for i in range(n):
            inp={**self.state.inputs}
            for key,svals in param_samples.items(): inp[key]=float(svals[i])
            try:
                res=Calculator(inp,mea,self.state.pol).calc_all()
                lcohs.append(min(res,key=lambda r:r["lcoh"])["lcoh"])
            except Exception: pass
        if not lcohs: self.mc_status_lbl.setText("No valid results."); return
        lcohs.sort(); nv=len(lcohs)
        mean_v=sum(lcohs)/nv; std_v=(sum((x-mean_v)**2 for x in lcohs)/nv)**0.5
        p10=lcohs[int(nv*0.10)]; p50=lcohs[int(nv*0.50)]; p90=lcohs[int(nv*0.90)]
        base_v=min(self.base_results,key=lambda r:r["lcoh"])["lcoh"]
        p_below=sum(1 for x in lcohs if x<base_v)/nv*100
        self.mc_status_lbl.setText(
            f"n={nv}   Mean: {mean_v:.3f}   Std: {std_v:.3f}\n"
            f"P10: {p10:.3f}   P50: {p50:.3f}   P90: {p90:.3f}\n"
            f"P(LCOH < base {base_v:.3f}): {p_below:.0f}%")
        # Histogram
        self.ax_hist.clear()
        if hasattr(self,'_cur_hist') and self._cur_hist:
            try: self._cur_hist.remove()
            except Exception: pass
        bins=25; mn_v=lcohs[max(0,int(nv*0.02))]; mx_v=lcohs[min(nv-1,int(nv*0.98))]
        # Filter first — prevents out-of-range samples piling into the last bin
        filtered_h=[v for v in lcohs if mn_v<=v<=mx_v]
        counts,bin_edges=np.histogram(filtered_h,bins=bins,range=(mn_v,mx_v))
        mids=[(bin_edges[i]+bin_edges[i+1])/2 for i in range(bins)]
        bw=float(bin_edges[1]-bin_edges[0]) if len(bin_edges)>1 else 1.0
        bar_colors=[BLUE_HEX if p10<=m<=p90 else BLUE_HEX+"55" for m in mids]
        bars=self.ax_hist.bar(mids,counts,width=bw*0.92,color=bar_colors,edgecolor="none")
        self.ax_hist.axvline(base_v,color=RED_HEX,linestyle="--",linewidth=1.0,alpha=0.7,label=f"Base {base_v:.3f}")
        self.ax_hist.axvspan(p10,p90,alpha=0.10,color=BLUE_HEX,label="P10–P90")
        self.ax_hist.set_xlabel("LCOH [$/kg H₂]",fontsize=9); self.ax_hist.set_ylabel("Count",fontsize=9)
        self.ax_hist.set_title(f"LCOH Distribution  (n={nv})",fontsize=10)
        self.ax_hist.legend(fontsize=7); self.ax_hist.grid(True,alpha=0.25)
        self._cur_hist=mplcursors.cursor(bars,hover=True)
        @self._cur_hist.connect("add")
        def _tt_hist(sel):
            sel.annotation.set_text(f"LCOH ≈ {sel.target[0]:.3f} $/kg\nCount = {int(sel.target[1])}")
            sel.annotation.get_bbox_patch().set(facecolor="white",alpha=0.9,edgecolor="#2D5A8E",linewidth=0.8)
            sel.annotation.set_fontsize(8)
            try:
                ax=sel.artist.axes; xl=ax.get_xlim()
                xf=(sel.target[0]-xl[0])/(xl[1]-xl[0]+1e-12)
                sel.annotation.xyann=(-70 if xf>0.85 else 0, -34)
            except Exception: sel.annotation.xyann=(0,-34)
            sel.annotation.set_anncoords("offset points")
        self.fig_hist.tight_layout(); self.canvas_hist.draw()
        # CDF
        self.ax_cdf.clear()
        if hasattr(self,'_cur_cdf') and self._cur_cdf:
            try: self._cur_cdf.remove()
            except Exception: pass
        step=max(1,nv//300)
        cx=[lcohs[i] for i in range(0,nv,step)]+[lcohs[-1]]
        cy=[i/nv*100 for i in range(0,nv,step)]+[100.0]
        cdf_line,=self.ax_cdf.plot(cx,cy,color=BLUE_HEX,linewidth=2)
        self.ax_cdf.axvline(base_v,color=RED_HEX,linestyle="--",linewidth=1.0,alpha=0.7,label=f"Base {base_v:.3f}")
        self.ax_cdf.set_xlabel("LCOH [$/kg H₂]",fontsize=9)
        self.ax_cdf.set_ylabel("Cumulative probability [%]",fontsize=9)
        self.ax_cdf.set_title("Cumulative Probability",fontsize=10)
        self.ax_cdf.set_ylim(0,100); self.ax_cdf.set_yticks([0,20,40,60,80,100])
        self.ax_cdf.legend(fontsize=7); self.ax_cdf.grid(True,alpha=0.25)
        self._cur_cdf=mplcursors.cursor(cdf_line,hover=True)
        @self._cur_cdf.connect("add")
        def _tt_cdf(sel):
            sel.annotation.set_text(f"LCOH = {sel.target[0]:.3f} $/kg\nP = {sel.target[1]:.1f}%")
            sel.annotation.get_bbox_patch().set(facecolor="white",alpha=0.9,edgecolor="#2D5A8E",linewidth=0.8)
            sel.annotation.set_fontsize(8); sel.annotation.xyann=_ann_offset(sel)
            sel.annotation.set_anncoords("offset points")
        self.fig_cdf.tight_layout(); self.canvas_cdf.draw()


# ── TAB 7: LCE ANALYSIS ──────────────────────────────────────────────────────

class LCETab(QWidget):
    def __init__(self, state, results_tab):
        super().__init__(); self.state=state; self.results_tab=results_tab
        self._non_elec=2.0935; self._spec_en=45.496   # (#9) internal only, not shown in UI
        self._plot_timer=QTimer(); self._plot_timer.setSingleShot(True)
        self._plot_timer.timeout.connect(self._update_plot)
        self._lce_hover_cid = None
        self._build()

    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(12,12,12,12)
        lay.addWidget(inst_label("Click 'Link from LCOH Results' to populate base values from the optimal operating point."))
        sp=QSplitter(Qt.Horizontal); lw=QWidget(); ll=QVBoxLayout(lw); ll.setSpacing(8)
        link_btn=QPushButton("Link from LCOH Results"); link_btn.setStyleSheet(BTN_PRIMARY)
        link_btn.clicked.connect(self._link); ll.addWidget(link_btn)

        grp_lnk=make_group("Parameters at Minimum LCOH")
        fl=QFormLayout(grp_lnk); fl.setSpacing(6)
        self.lbl_non_elec=QLabel("—"); fl.addRow("Non-electricity LCOH [$/kg]:",self.lbl_non_elec)
        self.lbl_spec_en =QLabel("—"); fl.addRow("Specific energy [kWh/kg H2]:",self.lbl_spec_en)
        self.lbl_elec_p  =QLabel("—"); fl.addRow("Electricity price [$/kWh]:",  self.lbl_elec_p)
        self.lbl_base    =QLabel("—"); fl.addRow("LCOH base [$/kg]:",            self.lbl_base)
        self.lbl_opt     =QLabel("—"); fl.addRow("Optimal operating point:",     self.lbl_opt)
        ll.addWidget(grp_lnk)

        # (#9) Manual override removed — user must link from LCOH Results

        grp_ref=make_group("Reference Parameters"); fr=QFormLayout(grp_ref); fr.setSpacing(6)
        self.sp_grey_lcoh=dspin(2.00,0,50,2); self.sp_grey_lce=dspin(12.0,0,200,1)
        self.sp_certifhy =dspin(4.4,0,20,1)
        fr.addRow("Grey H2 LCOH [$/kg]:",          self.sp_grey_lcoh)
        fr.addRow("Grey H2 LCE [kgCO2eq/kg H2]:",  self.sp_grey_lce)
        fr.addRow("CertifHy threshold [kgCO2eq/kg H2]:", self.sp_certifhy)
        ll.addWidget(grp_ref)

        grp_plt=make_group("Plot Settings"); fp=QFormLayout(grp_plt); fp.setSpacing(6)
        self.sp_cc_max  =dspin(200,1,2000,0); self.sp_lce_max=dspin(12.0,0.1,50,1)
        self.sp_vmin    =dspin(0.0,0,50,1);   self.sp_vmax   =dspin(0.0,0,50,1)
        self.sp_marker  =dspin(0.0,0,50,2);   self.sp_contour=dspin(0.0,0,5,2)
        fp.addRow("Carbon cost axis max [$/tCO2eq]:", self.sp_cc_max)
        fp.addRow("LCE axis max [kgCO2eq/kg H2]:",   self.sp_lce_max)
        fp.addRow("Colourbar min [$/kg]:",            self.sp_vmin)   # (#7) removed 0=auto
        fp.addRow("Colourbar max [$/kg]:",            self.sp_vmax)
        fp.addRow("AEMWE system LCE [kgCO2eq/kg H2]:", self.sp_marker)
        fp.addRow("Contour line gap [$/kg]:",          self.sp_contour)
        ll.addWidget(grp_plt); ll.addStretch()

        scroll=QScrollArea(); scroll.setWidgetResizable(True); scroll.setWidget(lw); sp.addWidget(scroll)
        rw=QWidget(); rl=QVBoxLayout(rw); rl.setContentsMargins(0,0,0,0)
        self.fig=Figure(); self.canvas=FigureCanvas(self.fig)
        self.canvas.setSizePolicy(QSizePolicy.Expanding,QSizePolicy.Expanding)
        rl.addLayout(chart_header(self,"Adjusted LCOH vs. Carbon Cost & LCE",self.fig,"LCE_analysis"))
        rl.addWidget(self.canvas); sp.addWidget(rw)
        sp.setSizes([260,810]); sp.setSizePolicy(QSizePolicy.Expanding,QSizePolicy.Expanding); lay.addWidget(sp,1)

        for w in [self.sp_grey_lcoh,self.sp_grey_lce,self.sp_certifhy,
                  self.sp_cc_max,self.sp_lce_max,self.sp_vmin,self.sp_vmax,self.sp_marker,self.sp_contour]:
            w.valueChanged.connect(lambda _: self._plot_timer.start(400))

    def _link(self):
        results=self.results_tab.results
        if not results: QMessageBox.warning(self,"No results","Run the LCOH calculation first."); return
        best=min(results,key=lambda r:r["lcoh"]); inp=self.state.inputs
        op_h=inp["operating_days_year"]*24.0
        spec_en=best["P_kw"]*op_h/best["h2_yr"]
        non_elec=best["lcoh"]-spec_en*inp["elec_price_kwh"]
        base=non_elec+spec_en*inp["elec_price_kwh"]
        self._non_elec=non_elec; self._spec_en=spec_en   # (#9) store internally
        self.lbl_non_elec.setText(f"{non_elec:.5f}"); self.lbl_spec_en.setText(f"{spec_en:.4f}")
        self.lbl_elec_p.setText(f"{inp['elec_price_kwh']:.5f}"); self.lbl_base.setText(f"{base:.5f}")
        self.lbl_opt.setText(f"j = {best['j']:.2f} A/cm2  -->  LCOH = {best['lcoh']:.4f} $/kg")
        grey_lce=self.sp_grey_lce.value(); grey_lcoh=self.sp_grey_lcoh.value(); certifhy=self.sp_certifhy.value()
        xmax=220.0
        if grey_lce>0 and (base-grey_lcoh)!=0:
            cc0=1000.0*(base-grey_lcoh)/grey_lce
            if cc0>0: xmax=max(xmax,cc0*1.4)
        if grey_lce>certifhy>0 and (base-grey_lcoh)!=0:
            cc_c=1000.0*(base-grey_lcoh)/(grey_lce-certifhy)
            if cc_c>0: xmax=max(xmax,cc_c*1.4)
        self.sp_cc_max.blockSignals(True); self.sp_cc_max.setValue(min(round(xmax/50)*50,1000)); self.sp_cc_max.blockSignals(False)
        self._update_plot()

    def _update_plot(self):
        from matplotlib.colors import Normalize as MNorm
        from matplotlib.gridspec import GridSpec
        # Disconnect stale hover handler
        if hasattr(self, '_lce_hover_cid') and self._lce_hover_cid:
            try: self.canvas.mpl_disconnect(self._lce_hover_cid)
            except Exception: pass
            self._lce_hover_cid = None

        self.fig.clf()
        gs  = GridSpec(1, 2, figure=self.fig, width_ratios=[20, 1], wspace=0.05)
        ax  = self.fig.add_subplot(gs[0, 0]); cax = self.fig.add_subplot(gs[0, 1])
        lcoh_base = self._non_elec + self._spec_en * self.state.inputs["elec_price_kwh"]
        grey_lcoh = self.sp_grey_lcoh.value(); grey_lce  = self.sp_grey_lce.value()
        certifhy  = self.sp_certifhy.value();  cc_max    = self.sp_cc_max.value()
        lce_max   = self.sp_lce_max.value()

        CC  = np.linspace(0, cc_max,  250); LCE = np.linspace(0, lce_max, 250)
        CC_g, LCE_g = np.meshgrid(CC, LCE)
        adj_green = lcoh_base + LCE_g * CC_g / 1000.0
        adj_grey  = grey_lcoh + grey_lce  * CC_g / 1000.0

        data_min = float(adj_green.min()); data_max = float(adj_green.max())
        rng = data_max - data_min
        auto_step = 0.1 if rng < 1 else (0.2 if rng < 3 else (0.5 if rng < 10 else 1.0))
        user_step = self.sp_contour.value(); step = user_step if user_step > 0 else auto_step
        vmin_sp = self.sp_vmin.value(); vmax_sp = self.sp_vmax.value()
        vmin_auto = np.floor(data_min / step) * step; vmax_auto = np.ceil(data_max / step) * step
        vmin = vmin_sp if vmin_sp > 0 else vmin_auto
        vmax = vmax_sp if vmax_sp > vmin else vmax_auto

        norm = MNorm(vmin=vmin, vmax=vmax)
        pm = ax.pcolormesh(CC_g, LCE_g, adj_green, cmap="viridis", norm=norm,
                           shading="gouraud", rasterized=True)

        # Set limits BEFORE clabel so labels are placed within visible area
        ax.set_xlim(0, cc_max); ax.set_ylim(0, lce_max)

        line_levels = np.arange(vmin_auto + step, vmax_auto, step)
        if len(line_levels) > 0:
            cs = ax.contour(CC_g, LCE_g, adj_green,
                            levels=line_levels, colors="black", linewidths=0.75, alpha=0.6)
            # Per-level valid-range label placement — works for any axis scale
            manual_positions = []
            for i, level in enumerate(line_levels):
                delta = level - lcoh_base
                if delta <= 0: continue
                # x range where label falls within LCE [3%..97%] of lce_max
                x_min_v = delta * 1000.0 / (lce_max * 0.97)
                x_max_v = delta * 1000.0 / (lce_max * 0.03)
                x_lo = max(cc_max * 0.03, x_min_v)
                x_hi = min(cc_max * 0.99, x_max_v)
                if x_lo >= x_hi: continue
                frac = 0.15 + 0.70 * (i / max(1, len(line_levels) - 1))
                x_pos = x_lo + frac * (x_hi - x_lo)
                lce_at = delta * 1000.0 / x_pos
                if 0 < lce_at < lce_max and 0 < x_pos < cc_max:
                    manual_positions.append((x_pos, lce_at))
            try:
                if manual_positions:
                    lbls = ax.clabel(cs, inline=True, fmt=lambda v: f"{v:.1f}",
                                     fontsize=7, manual=manual_positions, use_clabeltext=True)
                else:
                    lbls = ax.clabel(cs, inline=True, inline_spacing=4,
                                     fmt=lambda v: f"{v:.1f}", fontsize=7)
                    lbls = lbls or []
                for txt in (lbls or []): txt.set_color("black")
            except Exception: pass

        diff = adj_green - adj_grey
        SHADE_ALPHA = 0.30   # correct opacity

        # Separate grids sharing certifhy as EXACT boundary — no gap, no overlap
        EPS = lce_max * 0.05          # 5% buffer so contourf shading reaches y=0 after axis clipping
        n_w = max(3, int(200 * certifhy / max(lce_max, 1e-9)))
        LCE_bw = np.linspace(-EPS, certifhy, n_w)
        CC_bw, LCE_bw_g = np.meshgrid(CC, LCE_bw)
        diff_bw = (lcoh_base + LCE_bw_g * CC_bw / 1000.0) - (grey_lcoh + grey_lce * CC_bw / 1000.0)
        ax.contourf(CC_bw, LCE_bw_g, diff_bw, levels=[-1e10, 0], colors=["white"], alpha=SHADE_ALPHA)

        if 0 < certifhy < lce_max:
            n_r = max(3, int(200 * (lce_max - certifhy) / max(lce_max, 1e-9)))
            LCE_br = np.linspace(certifhy, lce_max, n_r)
            CC_br, LCE_br_g = np.meshgrid(CC, LCE_br)
            diff_br = (lcoh_base + LCE_br_g * CC_br / 1000.0) - (grey_lcoh + grey_lce * CC_br / 1000.0)
            ax.contourf(CC_br, LCE_br_g, diff_br, levels=[-1e10, 0], colors=[RED_HEX], alpha=SHADE_ALPHA)

        # Perimeter: analytical break-even curve — exact hyperbola, no grid artifacts
        # White below certifhy, red above; both share the same point at certifhy exactly
        def _be_line(lce_lo, lce_hi, n=250):
            lv = np.linspace(lce_lo, lce_hi, n)
            dv = grey_lce - lv
            with np.errstate(divide='ignore', invalid='ignore'):
                cv = np.where(np.abs(dv) > 1e-10, 1000.0*(lcoh_base-grey_lcoh)/dv, np.nan)
            cv = np.where((cv >= 0) & (cv <= cc_max), cv, np.nan)
            return cv, lv

        cc_w, lce_w = _be_line(0.0, min(certifhy, lce_max))
        ax.plot(cc_w, lce_w, 'w-', linewidth=1.4)

        if 0 < certifhy < lce_max:
            cc_r, lce_r = _be_line(certifhy, lce_max)
            ax.plot(cc_r, lce_r, '-', color=RED_HEX, linewidth=1.4)

        # Explicit horizontal bottom line at y=0 (perfectly straight)
        if grey_lce > 0 and (lcoh_base - grey_lcoh) != 0:
            cc_at_0 = 1000.0 * (lcoh_base - grey_lcoh) / grey_lce
            if 0 < cc_at_0 < cc_max:
                ax.plot([cc_at_0, cc_max], [0.0, 0.0], 'w-', linewidth=1.4, solid_capstyle='butt')

        # Invisible analytical line for hover — white portion only (below certifhy)
        cur_lce_max = min(certifhy if certifhy > 0 else lce_max, lce_max)
        lce_cur = np.linspace(0, cur_lce_max, 120)
        denom_cur = grey_lce - lce_cur
        with np.errstate(divide='ignore', invalid='ignore'):
            cc_cur = np.where(np.abs(denom_cur) > 1e-10,
                              1000.0 * (lcoh_base - grey_lcoh) / denom_cur, np.nan)
        cc_cur = np.where((cc_cur >= 0) & (cc_cur <= cc_max), cc_cur, np.nan)
        self._lce_ax = ax; self._lce_cc_cur = cc_cur
        self._lce_lce_cur = lce_cur; self._lce_base = lcoh_base

        # (#3) Annotations — positioned in clear space away from the curve, arrow points to curve
        ann_kw = dict(fontsize=9, fontweight="bold", color=RED_HEX, annotation_clip=True,
                      bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.85,
                                edgecolor=RED_HEX, linewidth=1.0))
        if grey_lce > 0 and (lcoh_base - grey_lcoh) != 0:
            cc0 = 1000.0 * (lcoh_base - grey_lcoh) / grey_lce
            if 0 < cc0 < cc_max * 0.95:
                # Text: left of curve, middle height — well clear of the perimeter
                txt_x = max(cc_max * 0.04, cc0 - cc_max * 0.28)
                txt_y = lce_max * 0.20
                ax.annotate(f"Break-even (LCE=0):\nCC = {cc0:.1f} $/tCO2",
                            xy=(cc0, 0), xytext=(txt_x, txt_y), ha="left",
                            **ann_kw, arrowprops=dict(arrowstyle="->", color=RED_HEX, lw=1.0))
        if 0 < certifhy < lce_max and grey_lce > certifhy:
            cc_c = 1000.0 * (lcoh_base - grey_lcoh) / (grey_lce - certifhy)
            if 0 < cc_c < cc_max * 0.95:
                txt_x = max(cc_max * 0.04, cc_c - cc_max * 0.28)
                txt_y = min(certifhy + lce_max * 0.22, lce_max * 0.88)
                ax.annotate(f"Break-even (CertifHy):\nCC = {cc_c:.1f} $/tCO2",
                            xy=(cc_c, certifhy), xytext=(txt_x, txt_y), ha="left",
                            **ann_kw, arrowprops=dict(arrowstyle="->", color=RED_HEX, lw=1.0))
        if 0 < certifhy < lce_max:
            ax.axhline(y=certifhy, color=RED_HEX, linewidth=0.9, linestyle="--")
            ax.text(cc_max * 0.02, certifhy + lce_max * 0.015,
                    f"CertifHy [{certifhy:.1f} kgCO2/kg]", fontsize=8, color=RED_HEX)
        lce_marker = self.sp_marker.value()
        if 0 < lce_marker < lce_max:
            ax.axhline(y=lce_marker, color="#F9A825", linewidth=1.0, linestyle="--")
            ax.text(cc_max * 0.02, lce_marker + lce_max * 0.015,
                    f"AEMWE LCE [{lce_marker:.2f} kgCO2/kg]", fontsize=8, color="#F9A825")

        cbar = self.fig.colorbar(pm, cax=cax)
        ticks = np.arange(vmin, vmax + step * 0.5, step)
        cbar.set_ticks(ticks); cbar.ax.set_yticklabels([f"{t:.1f}" for t in ticks])
        cbar.ax.tick_params(labelsize=8); cbar.set_label("Adjusted LCOH  [$ kg-1 H2]", fontsize=9)

        ax.set_xlabel("Carbon cost  [$/tCO2eq]", fontsize=10)
        ax.set_ylabel("Life cycle emissions  [kgCO2eq kg-1 H2]", fontsize=10)
        ax.set_xticks(np.linspace(0, cc_max, 6)); ax.set_yticks(np.linspace(0, lce_max, 5))
        ax.tick_params(axis="both", which="both", labelsize=9, width=0.75)
        ax.set_title(
            f"Adjusted LCOH vs. Carbon Cost & LCE\n"
            f"LCOH base = {lcoh_base:.3f} $/kg  |  Grey H2 LCOH = {grey_lcoh:.2f} $/kg"
            f" @ {grey_lce:.1f} kgCO2/kg", fontsize=10)

        try:
            self.fig.subplots_adjust(left=0.08, right=0.91, top=0.93, bottom=0.08, wspace=0.02)
        except Exception:
            pass
        self.canvas.draw()
        # Save background for blitting hover
        self._lce_bg = self.canvas.copy_from_bbox(ax.bbox)
        # Hidden annotation (created after draw so it's not in the baked background)
        self._lce_ann = ax.annotate("", xy=(0, 0), xytext=(14, 14),
            textcoords="offset points", visible=False,
            arrowprops=dict(arrowstyle="->", color="#444444", lw=0.8),
            bbox=dict(boxstyle="round,pad=0.25", facecolor="white", alpha=0.92,
                      edgecolor=BLUE_HEX, linewidth=0.8),
            fontsize=8, annotation_clip=False)
        self._lce_hover_cid = self.canvas.mpl_connect("motion_notify_event", self._on_lce_hover)

    def _on_lce_hover(self, event):
        """Blitting hover — only redraws the annotation, not the full figure."""
        ax = self._lce_ax; ann = self._lce_ann
        if event.inaxes != ax:
            if ann.get_visible():
                ann.set_visible(False)
                self.canvas.restore_region(self._lce_bg)
                self.canvas.blit(ax.bbox)
            return
        cc_cur = self._lce_cc_cur; lce_cur = self._lce_lce_cur
        valid = ~np.isnan(cc_cur)
        if not valid.any(): return
        cc_v = cc_cur[valid]; lce_v = lce_cur[valid]
        xlim = ax.get_xlim(); ylim = ax.get_ylim()
        dx = (cc_v - event.xdata)  / max(xlim[1] - xlim[0], 1e-9)
        dy = (lce_v - event.ydata) / max(ylim[1] - ylim[0], 1e-9)
        dist = np.sqrt(dx**2 + dy**2)
        idx = int(np.argmin(dist))
        if dist[idx] < 0.04:
            cc_val = cc_v[idx]; lce_val = lce_v[idx]
            lcoh_val = self._lce_base + lce_val * cc_val / 1000.0
            ann.set_text(
                f"Break-even\nCC = {cc_val:.1f} $/tCO2eq\n"
                f"LCE = {lce_val:.2f} kgCO2/kg\nLCOH = {lcoh_val:.3f} $/kg")
            ann.xy = (cc_val, lce_val); ann.set_visible(True)
            # smart offset for LCE annotation
            ax=self._lce_ax; xl=ax.get_xlim(); yl=ax.get_ylim()
            xf=(cc_val-xl[0])/(xl[1]-xl[0]+1e-12); yf=(lce_val-yl[0])/(yl[1]-yl[0]+1e-12)
            ann.xyann=((-70 if xf>0.85 else 14), (-34 if yf>0.78 else 14))
            self.canvas.restore_region(self._lce_bg)
            ax.draw_artist(ann)
            self.canvas.blit(ax.bbox)
        else:
            if ann.get_visible():
                ann.set_visible(False)
                self.canvas.restore_region(self._lce_bg)
                self.canvas.blit(ax.bbox)


# ── GUIDE TAB ────────────────────────────────────────────────────────────────

class GuideTab(QWidget):
    def __init__(self, state):
        super().__init__(); self.state=state; self._build()

    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(0, 0, 0, 0)
        self.browser = QTextBrowser(); self.browser.setOpenExternalLinks(False)
        self.browser.setStyleSheet("QTextBrowser { background-color: #F8F7F2; border: none; padding: 4px; color: #1A1A18; }")
        self.browser.setHtml(self._build_html()); lay.addWidget(self.browser)

    def _build_html(self):
        i=self.state.inputs
        T=i.get("operating_temp_c",60.0)
        vtn=Calculator.thermoneutral_voltage(T)
        return _GUIDE_HTML_STATIC + _build_equations_html(i, T, vtn)


def _build_equations_html(i, T_c, V_tn):
    """Generate the Key Equations section from live Calculator constants."""
    r=i.get("discount_rate_pct",8.0)/100.0; n=i.get("plant_lifetime_years",30.0)
    crf=(r*(1+r)**n/((1+r)**n-1)) if r>0 else 1/n
    return f"""
<h2 style="color:#2D5A8E; border-bottom:1px solid #C0BFB8; padding-bottom:6px;">Key Equations</h2>
<p style="font-size:10px;color:#888;">Generated from current inputs. Values shown in parentheses reflect the current parameter set.</p>
<table style="border-collapse:collapse; width:100%; margin-bottom:12px;">
  <tr style="vertical-align:top;">
    <td style="width:32px; padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">1</td>
    <td style="padding:7px 0; border-bottom:1px solid #C0BFB8;">
      <b>LCOH</b><br>
      <span style="font-family:monospace;font-size:12px;color:#1A5C1A;">LCOH = (Annualised CAPEX + Stack Replacement + OPEX) / H2 production [kg/yr]</span><br>
      <span style="color:#666;font-size:11px;">H2 production = {i.get('h2_target_kg_day',1000):.0f} kg/day × {i.get('operating_days_year',350):.0f} days/yr = {i.get('h2_target_kg_day',1000)*i.get('operating_days_year',350):,.0f} kg/yr</span>
    </td>
  </tr>
  <tr style="vertical-align:top;">
    <td style="padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">2</td>
    <td style="padding:7px 0; border-bottom:1px solid #C0BFB8;">
      <b>CAPEX</b><br>
      <span style="font-family:monospace;font-size:12px;color:#1A5C1A;">CAPEX = C_MEA + C_nonMEA + C_BoP</span><br>
      <span style="color:#666;font-size:11px;">
        C_nonMEA = {i.get('stack_non_mea_per_kw',80):.2f} $/kW_stack × stack power [kW]<br>
        C_BoP    = {i.get('bop_per_kw',300):.2f} $/kW_el × stack power [kW]
      </span>
    </td>
  </tr>
  <tr style="vertical-align:top;">
    <td style="padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">3</td>
    <td style="padding:7px 0; border-bottom:1px solid #C0BFB8;">
      <b>Capital Recovery Factor (CRF)</b><br>
      <span style="font-family:monospace;font-size:12px;color:#1A5C1A;">CRF = r×(1+r)^n / ((1+r)^n − 1)</span><br>
      <span style="color:#666;font-size:11px;">r = {i.get('discount_rate_pct',8):.1f}% | n = {i.get('plant_lifetime_years',30):.0f} years → CRF = <b>{crf:.6f}</b></span>
    </td>
  </tr>
  <tr style="vertical-align:top;">
    <td style="padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">4</td>
    <td style="padding:7px 0; border-bottom:1px solid #C0BFB8;">
      <b>Stack Replacement Cost (per year)</b><br>
      <span style="font-family:monospace;font-size:12px;color:#1A5C1A;">Stack Rep. = (op_hours/yr / t_life) × {i.get('stack_replacement_pct',40):.0f}% × (C_MEA + C_nonMEA)</span><br>
      <span style="color:#666;font-size:11px;">op_hours/yr = {i.get('operating_days_year',350):.0f} days × 24 = {i.get('operating_days_year',350)*24:.0f} h/yr</span>
    </td>
  </tr>
  <tr style="vertical-align:top;">
    <td style="padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">5</td>
    <td style="padding:7px 0; border-bottom:1px solid #C0BFB8;">
      <b>Stack Lifetime Degradation Model</b><br>
      <span style="font-family:monospace;font-size:12px;color:#1A5C1A;">t_life(j) = t_life(j_ref) × (j_ref / j)^n_deg</span><br>
      <span style="color:#666;font-size:11px;">j_ref = {i.get('j_ref',1):.2f} A/cm2 | t_life(j_ref) = {i.get('stack_lifetime_jref_h',20000):.0f} h | n_deg = {i.get('degradation_exp',1.5):.2f}</span>
    </td>
  </tr>
  <tr style="vertical-align:top;">
    <td style="padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">6</td>
    <td style="padding:7px 0; border-bottom:1px solid #C0BFB8;">
      <b>OPEX</b><br>
      <span style="font-family:monospace;font-size:12px;color:#1A5C1A;">OPEX = Electricity + Water + Maintenance + Labour</span><br>
      <span style="color:#666;font-size:11px;">
        Electricity  = stack power [kW] × {i.get('elec_price_kwh',0.034):.4f} $/kWh × op_hours/yr<br>
        Water        = H2 production [kg/yr] × {i.get('water_l_per_kg_h2',10):.1f} L/kg × {i.get('water_price_l',0.001):.4f} $/L<br>
        Maintenance  = {i.get('maintenance_pct_capex',2):.1f}% × CAPEX | Labour = {i.get('labour_yr',50000):,.0f} $/yr
      </span>
    </td>
  </tr>
  <tr style="vertical-align:top;">
    <td style="padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">7</td>
    <td style="padding:7px 0; border-bottom:1px solid #C0BFB8;">
      <b>Electrolyser Efficiency</b><br>
      <span style="font-family:monospace;font-size:12px;color:#1A5C1A;">η = V_tn(T) / V_cell × 100%</span><br>
      <span style="color:#666;font-size:11px;">
        V_tn(T) = (285830 − 48.5×(T−25)) / (2×96485)<br>
        At T = {T_c:.0f} °C → V_tn = <b>{V_tn:.5f} V</b>
      </span>
    </td>
  </tr>
  <tr style="vertical-align:top;">
    <td style="padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">8</td>
    <td style="padding:7px 0;">
      <b>LCE-Adjusted LCOH (LCE Analysis)</b><br>
      <span style="font-family:monospace;font-size:12px;color:#1A5C1A;">
        LCOH_adj(green) = LCOH_base + LCE_green × CC / 1000<br>
        LCOH_adj(grey)  = LCOH_grey  + LCE_grey  × CC / 1000
      </span><br>
      <span style="color:#666;font-size:11px;">CC = carbon cost [$/tCO2eq] | LCE in kgCO2eq/kg H2<br>
        CC_breakeven = 1000 × (LCOH_base − LCOH_grey) / (LCE_grey − LCE_green)
      </span>
    </td>
  </tr>
</table>
</body></html>
"""


_GUIDE_HTML_STATIC = """
<html>
<body style="font-family:'Segoe UI',Arial,sans-serif; font-size:12px;
             color:#1A1A18; background:#F8F7F2; margin:24px; line-height:1.7;">

<!-- ── ABOUT ──────────────────────────────────────────────────────────── -->
<p style="margin-top:0; margin-bottom:10px;">
  This calculator computes the economics of green hydrogen production from an
  Anion Exchange Membrane Water Electrolysis (AEMWE) system across a range of
  operating current densities, identifying the point that minimises hydrogen
  production cost.
</p>

<table style="border-collapse:collapse; width:100%; margin-bottom:16px;">
  <tr style="vertical-align:top;">
    <td style="width:32px; padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">1</td>
    <td style="padding:7px 0; border-bottom:1px solid #C0BFB8;">
      <b style="color:#1A1A18;">LCOH (Levelised Cost of Hydrogen)</b><br>
      The net present cost of producing one kilogram of hydrogen over the plant lifetime,
      accounting for capital expenditure (CAPEX), operating expenditure (OPEX), and the
      cost of capital. Expressed in <b>$/kg H2</b>.
    </td>
  </tr>
  <tr style="vertical-align:top;">
    <td style="padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">2</td>
    <td style="padding:7px 0; border-bottom:1px solid #C0BFB8;">
      <b style="color:#1A1A18;">LCE (Life Cycle Emissions)</b><br>
      Total greenhouse gas emissions associated with producing one kilogram of hydrogen,
      expressed in <b>kgCO2eq/kg H2</b>. The LCE Analysis tab shows how a carbon price
      affects the competitiveness of green hydrogen against grey (SMR) hydrogen, and
      identifies the break-even carbon cost.
    </td>
  </tr>
  <tr style="vertical-align:top;">
    <td style="padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">3</td>
    <td style="padding:7px 0;">
      <b style="color:#1A1A18;">Sensitivity Analysis</b><br>
      A systematic assessment of how much the LCOH changes when individual input parameters
      are varied by &plusmn;20% from the base case, showing which parameters drive LCOH most.
    </td>
  </tr>
</table>

<!-- ── WORKFLOW ────────────────────────────────────────────────────────── -->
<h2 style="color:#2D5A8E; border-bottom:1px solid #C0BFB8; padding-bottom:6px;">
  Workflow
</h2>
<table style="border-collapse:collapse; width:100%; margin-bottom:12px;">
  <tr style="vertical-align:top;">
    <td style="width:32px; padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">1</td>
    <td style="padding:7px 0; border-bottom:1px solid #C0BFB8;">
      <b style="color:#1A1A18;">Inputs</b><br>
      Set plant design targets, financial parameters (lifetime, discount rate),
      electrochemical constants, capital and operating expenditures, and the
      AEMWE system polarisation curve. Click <b>Apply Changes</b> when done.
    </td>
  </tr>
  <tr style="vertical-align:top;">
    <td style="padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">2</td>
    <td style="padding:7px 0; border-bottom:1px solid #C0BFB8;">
      <b style="color:#1A1A18;">BOM / MEA Materials</b><br>
      Select materials and quantities for each MEA component. Override reference
      costs with custom supplier quotes using the Override checkbox. Click
      <b>Apply Changes</b> when done.
    </td>
  </tr>
  <tr style="vertical-align:top;">
    <td style="padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">3</td>
    <td style="padding:7px 0; border-bottom:1px solid #C0BFB8;">
      <b style="color:#1A1A18;">Anode / Cathode Catalyst Synthesis Cost</b><br>
      For custom catalyst synthesis only. To unlock these tabs, first set the
      <b>Anode Catalyst</b> and/or <b>Cathode Catalyst</b> in <b>BOM / MEA Materials</b>
      to <b>&ldquo;Custom&rdquo;</b>. Then enter the synthesis route,
      yield, reagents, and solvents. Click <b>Apply Changes</b> to push the computed
      cost back into the BOM.
    </td>
  </tr>
  <tr style="vertical-align:top;">
    <td style="padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">4</td>
    <td style="padding:7px 0; border-bottom:1px solid #C0BFB8;">
      <b style="color:#1A1A18;">LCOH Results</b><br>
      Click <b>Calculate LCOH</b> to run the model across all polarisation curve
      operating points. The table and charts update automatically. Use the Plot
      Settings spinboxes to zoom the y-axis. Use <b>Export</b> to save results as
      CSV, TXT, JSON, or XLSX.
    </td>
  </tr>
  <tr style="vertical-align:top;">
    <td style="padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">5</td>
    <td style="padding:7px 0; border-bottom:1px solid #C0BFB8;">
      <b style="color:#1A1A18;">LCE Analysis</b><br>
      Click <b>Link from LCOH Results</b> to populate the base LCOH and specific
      energy from the optimal operating point. Set the grey hydrogen LCOH and LCE
      reference values, then adjust the carbon cost axis to visualise break-even
      carbon pricing for both LCE = 0 and the CertifHy threshold.
    </td>
  </tr>
  <tr style="vertical-align:top;">
    <td style="padding:7px 10px 7px 0; color:#B03030; font-weight:bold; font-size:14px;">6</td>
    <td style="padding:7px 0;">
      <b style="color:#1A1A18;">Sensitivity Analysis</b><br>
      Click <b>Set Base Case</b> to lock the reference. Use the sliders to explore
      &plusmn;20% deviations in key parameters. For probabilistic analysis, set the
      distribution and uncertainty ranges under Monte Carlo Analysis, then click
      <b>Run Monte Carlo</b>.
    </td>
  </tr>
</table>

<!-- ── KEY EQUATIONS generated programmatically ──────────────────────── -->
"""


# ── MAIN WINDOW ───────────────────────────────────────────────────────────────

# Status pill style — matches BTN_PRIMARY font/padding exactly
_PILL_OK  = ("QPushButton { background:#2E7D32; color:white; border-radius:4px;"
             " padding:4px 12px; font-size:9px; font-weight:bold; border:none; }"
             "QPushButton:hover { background:#2E7D32; }")
_PILL_WARN= ("QPushButton { background:#E65100; color:white; border-radius:4px;"
             " padding:4px 12px; font-size:9px; font-weight:bold; border:none; }"
             "QPushButton:hover { background:#E65100; }")
# Save/Load match pill exactly (same font-size/padding, different hue)
_BTN_SB   = ("QPushButton { background:#2D5A8E; color:white; border-radius:4px;"
             " padding:4px 12px; font-size:9px; font-weight:bold; border:none; }"
             "QPushButton:hover  { background:#3B6BA0; }"
             "QPushButton:pressed{ background:#1E4070; }")

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("LCOH Calculator for AEMWE  |  Ananta Fareza")
        try:
            icon=QFileIconProvider().icon(QFileInfo(sys.executable))
            if not icon.isNull(): self.setWindowIcon(icon)
        except Exception: pass

        state=AppState()
        self.state=state
        self.tabs=QTabWidget()

        self.guide_tab        = GuideTab(state)
        self.inputs_tab       = InputsTab(state)
        self.bom_tab          = BOMTab(state)
        self.anode_cat_tab    = CatalystSynthesisTab(state, "Anode")
        self.cathode_cat_tab  = CatalystSynthesisTab(state, "Cathode")
        self.results_tab      = ResultsTab(state)
        self.sens_tab         = SensitivityTab(state)
        self.lce_tab          = LCETab(state, self.results_tab)

        self.tabs.addTab(self.guide_tab,       "Guide")
        self.tabs.addTab(self.inputs_tab,      "Inputs")
        self.tabs.addTab(self.anode_cat_tab,   "Anode Catalyst Synthesis Cost")
        self.tabs.addTab(self.cathode_cat_tab, "Cathode Catalyst Synthesis Cost")
        self.tabs.addTab(self.bom_tab,         "BOM / MEA Materials")
        self.tabs.addTab(self.results_tab,     "LCOH Results")
        self.tabs.addTab(self.lce_tab,         "LCE Analysis")
        self.tabs.addTab(self.sens_tab,        "Sensitivity Analysis")

        self._anode_tab_idx   = self.tabs.indexOf(self.anode_cat_tab)
        self._cathode_tab_idx = self.tabs.indexOf(self.cathode_cat_tab)

        self.bom_tab._combos["Anode Catalyst"].currentTextChanged.connect(self._update_synth_tabs)
        self.bom_tab._combos["Cathode Catalyst"].currentTextChanged.connect(self._update_synth_tabs)
        self._update_synth_tabs()

        # Wire dirty callbacks
        self.inputs_tab.set_dirty_callback(self._inputs_changed)
        self.bom_tab.set_dirty_callback(self._bom_changed)

        # Patch calculate to update pills
        orig_calc=self.results_tab.calculate
        def _patched_calc():
            orig_calc()
            if self.results_tab.results:
                self._pill_results.setText("LCOH: ready"); self._pill_results.setStyleSheet(_PILL_OK)
                self._pill_lce.setText("LCE: outdated"); self._pill_lce.setStyleSheet(_PILL_WARN)
        self.results_tab.calculate=_patched_calc

        orig_link=self.lce_tab._link
        def _patched_link():
            orig_link()
            self._pill_lce.setText("LCE: ready"); self._pill_lce.setStyleSheet(_PILL_OK)
        self.lce_tab._link=_patched_link

        # Central widget — tabs only
        self.setCentralWidget(self.tabs)

        # ── Status bar ────────────────────────────────────────────────────
        # Left: regular showMessage tooltip (restored)
        # Right: pills + Save/Load via addPermanentWidget
        sb=QStatusBar(); self.setStatusBar(sb)
        sb.setSizeGripEnabled(True)
        sb.showMessage("Ready — apply your inputs, then switch to LCOH Results and click Calculate LCOH.")

        # Pills — non-interactive, right-aligned via addPermanentWidget
        self._pill_inputs =QPushButton("Inputs: applied")
        self._pill_bom    =QPushButton("BOM: applied")
        self._pill_results=QPushButton("LCOH: ready")
        self._pill_lce    =QPushButton("LCE: ready")
        for p in [self._pill_inputs,self._pill_bom,self._pill_results,self._pill_lce]:
            p.setStyleSheet(_PILL_OK); p.setFocusPolicy(Qt.NoFocus)

        sep=QFrame(); sep.setFrameShape(QFrame.VLine); sep.setFrameShadow(QFrame.Sunken)

        save_btn=QPushButton("Save Project"); save_btn.setStyleSheet(_BTN_SB)
        load_btn=QPushButton("Load Project"); load_btn.setStyleSheet(_BTN_SB)
        save_btn.clicked.connect(self._save_project)
        load_btn.clicked.connect(self._load_project)

        # addPermanentWidget puts items on the RIGHT and they don't get covered by showMessage
        for w in [self._pill_inputs, self._pill_bom, self._pill_results, self._pill_lce,
                  sep, save_btn, load_btn]:
            sb.addPermanentWidget(w)

    def _inputs_changed(self, clean=False):
        if clean:
            self._pill_inputs.setText("Inputs: applied"); self._pill_inputs.setStyleSheet(_PILL_OK)
        else:
            self._pill_inputs.setText("Inputs: modified"); self._pill_inputs.setStyleSheet(_PILL_WARN)
        self._pill_results.setText("LCOH: outdated"); self._pill_results.setStyleSheet(_PILL_WARN)
        self._pill_lce.setText("LCE: outdated");      self._pill_lce.setStyleSheet(_PILL_WARN)

    def _bom_changed(self, clean=False):
        if clean:
            self._pill_bom.setText("BOM: applied"); self._pill_bom.setStyleSheet(_PILL_OK)
        else:
            self._pill_bom.setText("BOM: modified"); self._pill_bom.setStyleSheet(_PILL_WARN)
        self._pill_results.setText("LCOH: outdated"); self._pill_results.setStyleSheet(_PILL_WARN)
        self._pill_lce.setText("LCE: outdated");      self._pill_lce.setStyleSheet(_PILL_WARN)

    def _update_synth_tabs(self):
        anode_custom   = self.bom_tab._combos["Anode Catalyst"].currentText() == "Custom"
        cathode_custom = self.bom_tab._combos["Cathode Catalyst"].currentText() == "Custom"
        self.tabs.setTabEnabled(self._anode_tab_idx,   anode_custom)
        self.tabs.setTabEnabled(self._cathode_tab_idx, cathode_custom)

    # ── Save / Load ───────────────────────────────────────────────────────
    def _save_project(self):
        path,_=QFileDialog.getSaveFileName(self,"Save Project","LCOH_project","LCOH Project (*.lcoh);;JSON (*.json)")
        if not path: return
        if not (path.endswith(".lcoh") or path.endswith(".json")): path+=".lcoh"
        try:
            with open(path,"w") as f: json.dump(self.state.to_dict(),f,indent=2)
            QMessageBox.information(self,"Saved",f"Project saved:\n{path}")
        except Exception as ex:
            QMessageBox.critical(self,"Save error",str(ex))

    def _load_project(self):
        path,_=QFileDialog.getOpenFileName(self,"Load Project","","LCOH Project (*.lcoh);;JSON (*.json);;All files (*)")
        if not path: return
        try:
            with open(path,"r") as f: d=json.load(f)
        except Exception as ex:
            QMessageBox.critical(self,"Load error",f"Could not read file:\n{ex}"); return
        warnings=self.state.from_dict(d)
        # Repopulate all tabs
        self.inputs_tab.load_state()
        self.bom_tab.load_state()
        self._update_synth_tabs()
        self.results_tab.results=[]
        # Reset status pills
        self._pill_inputs.setText("Inputs: applied");    self._pill_inputs.setStyleSheet(_PILL_OK)
        self._pill_bom.setText("BOM: applied");          self._pill_bom.setStyleSheet(_PILL_OK)
        self._pill_results.setText("LCOH: outdated");    self._pill_results.setStyleSheet(_PILL_WARN)
        self._pill_lce.setText("LCE: outdated");         self._pill_lce.setStyleSheet(_PILL_WARN)
        msg=f"Project loaded from:\n{path}"
        if warnings: msg+="\n\nWarnings:\n"+"\n".join(f"• {w}" for w in warnings)
        QMessageBox.information(self,"Loaded",msg)


# ── ENTRY POINT ───────────────────────────────────────────────────────────────

def main():
    app=QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setPalette(make_palette())
    f=QFont("Segoe UI",9) if sys.platform=="win32" else QFont("SF Pro Text",10)
    app.setFont(f)
    win=MainWindow(); win.show(); win.showMaximized()
    sys.exit(app.exec())

if __name__=="__main__":
    main()